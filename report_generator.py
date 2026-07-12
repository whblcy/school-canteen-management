"""
报表生成器 - 基于模板导出各类食堂报表
"""
import os
import sys
import shutil
from datetime import datetime, timedelta
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import QFileDialog, QMessageBox

from database import (CategoryDAO, SupplierDAO, IngredientDAO,
                      StockInDAO, StockOutDAO, InventoryCheckDAO,
                      ReportDAO, InspectionRecordDAO, get_connection)


def get_template_dir():
    """获取模板目录路径"""
    if hasattr(sys, '_MEIPASS'):
        base_dir = sys._MEIPASS
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, '表格')


def get_template_path(filename):
    """获取模板文件完整路径"""
    return os.path.join(get_template_dir(), filename)


def copy_cell_style(src_cell, dst_cell):
    """复制单元格样式"""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def insert_row_with_style(ws, row_idx, max_col):
    """在指定行插入新行，并复制上一行的样式"""
    ws.insert_rows(row_idx)
    for col in range(1, max_col + 1):
        src_cell = ws.cell(row=row_idx + 1, column=col)
        dst_cell = ws.cell(row=row_idx, column=col)
        copy_cell_style(src_cell, dst_cell)


def copy_sheet_structure(src_ws, dst_ws):
    """复制整个sheet的结构（合并单元格、列宽、行高、样式等）"""
    dst_ws.sheet_format = copy(src_ws.sheet_format)
    dst_ws.sheet_properties = copy(src_ws.sheet_properties)
    dst_ws.page_margins = copy(src_ws.page_margins)
    dst_ws.page_setup = copy(src_ws.page_setup)
    
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter] = copy(col_dim)
    
    for row_idx, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx] = copy(row_dim)
    
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))
    
    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column)
            copy_cell_style(cell, dst_cell)


def get_stock_in_by_date(date_str):
    """按日期查询入库记录"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT c.name as category_name, i.name as ingredient_name, i.unit,
                   SUM(si.quantity) as total_quantity,
                   SUM(si.total_price) / SUM(si.quantity) as avg_price,
                   SUM(si.total_price) as total_amount
            FROM stock_in si
            JOIN ingredients i ON si.ingredient_id = i.id
            LEFT JOIN categories c ON i.category_id = c.id
            WHERE DATE(si.created_at) = ?
            GROUP BY si.ingredient_id
            ORDER BY c.name, i.name
        ''', (date_str,))
        return [dict(row) for row in cursor.fetchall()]


def get_stock_out_by_date(date_str):
    """按日期查询出库记录"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT c.name as category_name, i.name as ingredient_name, i.unit,
                   SUM(so.quantity) as total_quantity,
                   SUM(so.total_price) / SUM(so.quantity) as avg_price,
                   SUM(so.total_price) as total_amount
            FROM stock_out so
            JOIN ingredients i ON so.ingredient_id = i.id
            LEFT JOIN categories c ON i.category_id = c.id
            WHERE DATE(so.created_at) = ?
            GROUP BY so.ingredient_id
            ORDER BY c.name, i.name
        ''', (date_str,))
        return [dict(row) for row in cursor.fetchall()]


def get_days_in_month(year, month):
    """获取某月的天数"""
    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)
    last_day = next_month - timedelta(days=1)
    return last_day.day


class ReportGenerator:
    """报表生成器"""

    @staticmethod
    def export_daily_stock_sheet(parent_widget, year=None, month=None, file_path=None):
        """导出每日出入库表"""
        if year is None:
            year = datetime.now().year
        if month is None:
            month = datetime.now().month

        if file_path is None:
            file_path, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "导出每日出入库表",
                f"每日出入库表_{year}年{month}月.xlsx",
                "Excel Files (*.xlsx)"
            )

        if not file_path:
            return False

        try:
            template_path = get_template_path('2026年6月份每天出入库表.xlsx')
            if not os.path.exists(template_path):
                QMessageBox.critical(parent_widget, "模板不存在", f"找不到模板文件:\n{template_path}")
                return False

            shutil.copy2(template_path, file_path)
            wb = load_workbook(file_path)
            template_ws = wb['出入库表']

            days = get_days_in_month(year, month)
            template_block_height = 28

            for day in range(days, 0, -1):
                date_str = f"{year}-{month:02d}-{day:02d}"
                date_obj = datetime(year, month, day)
                
                src_start_row = 1
                src_end_row = 27
                
                new_sheet_name = f"{month}月{day}日"
                
                new_ws = wb.copy_worksheet(template_ws)
                new_ws.title = new_sheet_name
                
                date_cell = new_ws.cell(row=2, column=4)
                date_cell.value = date_obj
                
                in_data = get_stock_in_by_date(date_str)
                out_data = get_stock_out_by_date(date_str)
                
                in_start_row = 5
                in_template_row = 5
                max_in_rows = 19
                
                if len(in_data) > max_in_rows:
                    for i in range(len(in_data) - max_in_rows):
                        insert_row_with_style(new_ws, in_start_row + max_in_rows, 7)
                
                for idx, item in enumerate(in_data):
                    row = in_start_row + idx
                    new_ws.cell(row=row, column=1, value=item.get('category_name', ''))
                    new_ws.cell(row=row, column=2, value=item.get('ingredient_name', ''))
                    new_ws.cell(row=row, column=3, value=item.get('total_quantity', 0))
                    new_ws.cell(row=row, column=4, value=item.get('unit', ''))
                    avg_price = item.get('avg_price', 0) or 0
                    new_ws.cell(row=row, column=5, value=round(avg_price, 2))
                    new_ws.cell(row=row, column=6, value=item.get('total_amount', 0))
                
                out_start_row = 5
                max_out_rows = 19
                
                if len(out_data) > max_out_rows:
                    for i in range(len(out_data) - max_out_rows):
                        insert_row_with_style(new_ws, out_start_row + max_out_rows, 16)
                
                for idx, item in enumerate(out_data):
                    row = out_start_row + idx
                    new_ws.cell(row=row, column=9, value=item.get('category_name', ''))
                    new_ws.cell(row=row, column=10, value=item.get('ingredient_name', ''))
                    new_ws.cell(row=row, column=11, value=item.get('unit', ''))
                    qty = item.get('total_quantity', 0)
                    new_ws.cell(row=row, column=12, value=qty)
                    new_ws.cell(row=row, column=13, value=qty)
                    avg_price = item.get('avg_price', 0) or 0
                    new_ws.cell(row=row, column=14, value=round(avg_price, 2))
                    new_ws.cell(row=row, column=15, value=item.get('total_amount', 0))

            if '出入库表' in wb.sheetnames:
                del wb['出入库表']

            wb.save(file_path)
            wb.close()
            return True

        except Exception as e:
            QMessageBox.critical(parent_widget, "导出失败", f"导出时发生错误:\n{str(e)}")
            return False

    @staticmethod
    def export_monthly_summary(parent_widget, year=None, month=None, file_path=None):
        """导出每月出入库统计表"""
        if year is None:
            year = datetime.now().year
        if month is None:
            month = datetime.now().month

        if file_path is None:
            file_path, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "导出每月汇总表",
                f"每月出入库统计表_{year}年{month}月.xlsx",
                "Excel Files (*.xlsx)"
            )

        if not file_path:
            return False

        try:
            template_path = get_template_path('2026年6月每月出入库汇总等表和14、每月结算食材公示.xlsx')
            if not os.path.exists(template_path):
                QMessageBox.critical(parent_widget, "模板不存在", f"找不到模板文件:\n{template_path}")
                return False

            shutil.copy2(template_path, file_path)
            wb = load_workbook(file_path)
            
            ws = wb['5.每月出入库统计表']
            
            ws.cell(row=1, column=1, value=f"{year}年{month}月营养餐食材出入库统计表")
            days = get_days_in_month(year, month)
            ws.cell(row=2, column=9, value=f"{year}年{month}月1日至{days}日")
            
            stock_in_data = ReportDAO.get_monthly_in(year, month)
            stock_out_data = ReportDAO.get_monthly_out(year, month)
            all_ingredients = IngredientDAO.get_all()
            categories = {cat.id: cat.name for cat in CategoryDAO.get_all()}
            
            in_map = {item['ingredient_name']: item for item in stock_in_data}
            out_map = {item['ingredient_name']: item for item in stock_out_data}
            
            data_rows = []
            for ing in all_ingredients:
                cat_name = categories.get(ing.category_id, '')
                in_item = in_map.get(ing.name, {'total_quantity': 0, 'total_amount': 0})
                out_item = out_map.get(ing.name, {'total_quantity': 0, 'total_amount': 0})
                data_rows.append({
                    'category': cat_name,
                    'name': ing.name,
                    'unit': ing.unit,
                    'in_qty': in_item['total_quantity'],
                    'in_amount': in_item['total_amount'],
                    'out_qty': out_item['total_quantity'],
                    'out_amount': out_item['total_amount'],
                })
            
            data_start_row = 5
            template_last_data_row = 64
            max_data_rows = template_last_data_row - data_start_row + 1
            
            if len(data_rows) > max_data_rows:
                for i in range(len(data_rows) - max_data_rows):
                    insert_row_with_style(ws, template_last_data_row + i, 15)
            
            for idx, row_data in enumerate(data_rows):
                row = data_start_row + idx
                ws.cell(row=row, column=1, value=row_data['category'])
                ws.cell(row=row, column=2, value=row_data['name'])
                ws.cell(row=row, column=3, value=row_data['unit'])
                ws.cell(row=row, column=4, value=0)
                ws.cell(row=row, column=5, value=0)
                ws.cell(row=row, column=6, value=row_data['unit'])
                ws.cell(row=row, column=7, value=row_data['in_qty'])
                ws.cell(row=row, column=8, value=row_data['in_amount'])
                ws.cell(row=row, column=9, value=row_data['unit'])
                ws.cell(row=row, column=10, value=row_data['out_qty'])
                ws.cell(row=row, column=11, value=row_data['out_amount'])
                ws.cell(row=row, column=12, value=row_data['unit'])
                ws.cell(row=row, column=13, value=f"=D{row}+G{row}-J{row}")
                ws.cell(row=row, column=14, value=f"=E{row}+H{row}-K{row}")
            
            wb.save(file_path)
            wb.close()
            return True

        except Exception as e:
            QMessageBox.critical(parent_widget, "导出失败", f"导出时发生错误:\n{str(e)}")
            return False

    @staticmethod
    def export_financial_report(parent_widget, year=None, file_path=None):
        """导出财务收支情况表"""
        if year is None:
            year = datetime.now().year

        if file_path is None:
            file_path, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "导出财务收支情况表",
                f"财务收支情况表_{year}年度.xlsx",
                "Excel Files (*.xlsx)"
            )

        if not file_path:
            return False

        try:
            template_path = get_template_path('附件1 ：月份、年度食堂（营养餐）财务收支情况表(1).xlsx')
            if not os.path.exists(template_path):
                QMessageBox.critical(parent_widget, "模板不存在", f"找不到模板文件:\n{template_path}")
                return False

            shutil.copy2(template_path, file_path)
            wb = load_workbook(file_path)
            
            category_name_map = {
                '肉禽奶蛋类': ['肉类', '蛋类', '水产类'],
                '蔬菜瓜果类': ['蔬菜类', '水果类'],
                '粮食类': ['粮油类'],
                '干货类': [],
                '食用油类': [],
                '食品类': ['豆制品'],
            }
            
            for month in range(1, 13):
                sheet_name = None
                for name in wb.sheetnames:
                    if str(month) in name and '月' in name:
                        sheet_name = name
                        break
                
                if sheet_name is None:
                    continue
                
                ws = wb[sheet_name]
                
                ws.cell(row=1, column=1, value=f"附件1 {month}月份、年度食堂(营养餐）财务收支情况表")
                ws.cell(row=2, column=1, value=f"编制单位：容县松山镇中心学校    {year}年{month}月")
                
                finance_data = ReportDAO.get_monthly_finance(year, month)
                category_in = {item['category_name']: item['amount'] for item in finance_data['category_in']}
                
                expense_map = {}
                for expense_cat, source_cats in category_name_map.items():
                    total = 0
                    for src_cat in source_cats:
                        total += category_in.get(src_cat, 0)
                    expense_map[expense_cat] = total
                
                ws.cell(row=6, column=5, value=expense_map.get('肉禽奶蛋类', 0))
                ws.cell(row=7, column=5, value=expense_map.get('蔬菜瓜果类', 0))
                ws.cell(row=8, column=5, value=expense_map.get('粮食类', 0))
                ws.cell(row=9, column=5, value=expense_map.get('干货类', 0))
                ws.cell(row=10, column=5, value=expense_map.get('食用油类', 0))
                ws.cell(row=11, column=5, value=expense_map.get('食品类', 0))
            
            wb.save(file_path)
            wb.close()
            return True

        except Exception as e:
            QMessageBox.critical(parent_widget, "导出失败", f"导出时发生错误:\n{str(e)}")
            return False

    @staticmethod
    def export_inventory_check_sheet(parent_widget, file_path=None):
        """导出库存物品盘存盘亏表"""
        if file_path is None:
            file_path, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "导出库存盘存盘亏表",
                f"库存物品盘存盘亏表_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)"
            )

        if not file_path:
            return False

        try:
            template_path = get_template_path('库存物品盘存盘亏表.xlsx')
            if not os.path.exists(template_path):
                QMessageBox.critical(parent_widget, "模板不存在", f"找不到模板文件:\n{template_path}")
                return False

            shutil.copy2(template_path, file_path)
            wb = load_workbook(file_path)
            ws = wb['6.库存物品盘存盘亏表']
            
            now = datetime.now()
            ws.cell(row=2, column=6, value=f"  {now.year} 年 {now.month} 月 {now.day} 日编制")
            
            all_ingredients = IngredientDAO.get_all()
            categories = {cat.id: cat.name for cat in CategoryDAO.get_all()}
            
            category_order = [
                '粮食类', '粮油类', '蔬菜类', '肉类', '蛋类', '水产类',
                '调味品', '豆制品', '水果类'
            ]
            
            display_categories = [
                ('1.粮食', ['粮油类']),
                ('2.蔬菜', ['蔬菜类']),
                ('3.调料', ['调味品']),
                ('4.肉食品', ['肉类']),
                ('5.水产品', ['水产类']),
                ('6.蛋奶品', ['蛋类']),
                ('7.食用油', []),
                ('8.其他', ['豆制品', '水果类']),
            ]
            
            checks = InventoryCheckDAO.get_all(1000)
            check_map = {c.ingredient_id: c for c in checks}
            
            data_rows = []
            for disp_name, cat_names in display_categories:
                data_rows.append({'is_category': True, 'name': disp_name})
                for ing in all_ingredients:
                    cat_name = categories.get(ing.category_id, '')
                    if cat_name in cat_names:
                        check = check_map.get(ing.id)
                        if check:
                            system_stock = check.system_stock
                            actual_stock = check.actual_stock
                            difference = check.difference
                        else:
                            system_stock = ing.current_stock
                            actual_stock = ing.current_stock
                            difference = 0
                        
                        avg_price = 0
                        with get_connection() as conn:
                            cursor = conn.cursor()
                            cursor.execute('''
                                SELECT unit_price FROM stock_in 
                                WHERE ingredient_id = ? 
                                ORDER BY created_at DESC LIMIT 1
                            ''', (ing.id,))
                            row = cursor.fetchone()
                            if row:
                                avg_price = row['unit_price']
                        
                        data_rows.append({
                            'is_category': False,
                            'name': ing.name,
                            'unit': ing.unit,
                            'system_stock': system_stock,
                            'actual_stock': actual_stock,
                            'difference': difference,
                            'avg_price': avg_price,
                            'amount': avg_price * difference,
                            'remark': check.remark if check else '',
                        })
            
            data_start_row = 4
            template_last_row = 29
            
            if len(data_rows) > (template_last_row - data_start_row + 1):
                extra_rows = len(data_rows) - (template_last_row - data_start_row + 1)
                for i in range(extra_rows):
                    insert_row_with_style(ws, template_last_row + i, 8)
            
            for idx, row_data in enumerate(data_rows):
                row = data_start_row + idx
                if row_data['is_category']:
                    ws.cell(row=row, column=1, value=row_data['name'])
                else:
                    ws.cell(row=row, column=1, value=row_data['name'])
                    ws.cell(row=row, column=2, value=row_data['unit'])
                    ws.cell(row=row, column=3, value=row_data['system_stock'])
                    ws.cell(row=row, column=4, value=row_data['actual_stock'])
                    ws.cell(row=row, column=5, value=row_data['difference'])
                    ws.cell(row=row, column=6, value=row_data['avg_price'])
                    ws.cell(row=row, column=7, value=row_data['amount'])
                    ws.cell(row=row, column=8, value=row_data['remark'])
            
            wb.save(file_path)
            wb.close()
            return True

        except Exception as e:
            QMessageBox.critical(parent_widget, "导出失败", f"导出时发生错误:\n{str(e)}")
            return False

    @staticmethod
    def export_inspection_report(parent_widget, file_path=None):
        """导出进货查验记录表"""
        if file_path is None:
            file_path, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "导出进货查验记录表",
                f"进货查验记录表_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)"
            )

        if not file_path:
            return False

        try:
            template_path = get_template_path('进货查验记录表.xlsx')
            if not os.path.exists(template_path):
                QMessageBox.critical(parent_widget, "模板不存在", f"找不到模板文件:\n{template_path}")
                return False

            shutil.copy2(template_path, file_path)
            wb = load_workbook(file_path)
            ws = wb['Sheet1']
            
            records = InspectionRecordDAO.get_all(1000)
            
            data_start_row = 4
            template_last_row = 1006
            max_data_rows = template_last_row - data_start_row + 1
            
            if len(records) > max_data_rows:
                extra_rows = len(records) - max_data_rows
                for i in range(extra_rows):
                    insert_row_with_style(ws, template_last_row + i, 16)
            
            ingredients = {ing.id: ing for ing in IngredientDAO.get_all()}
            suppliers = {s.id: s for s in SupplierDAO.get_all()}
            
            current_date = None
            for idx, record in enumerate(records):
                row = data_start_row + idx
                
                ing = ingredients.get(record.ingredient_id)
                supplier = None
                if ing and ing.supplier_id:
                    supplier = suppliers.get(ing.supplier_id)
                
                record_date = record.production_date or record.inspection_date or record.created_at[:10]
                if record_date != current_date:
                    ws.cell(row=row, column=1, value=record_date)
                    current_date = record_date
                
                ws.cell(row=row, column=2, value=record.ingredient_name)
                ws.cell(row=row, column=3, value=record.unit)
                ws.cell(row=row, column=4, value=record.quantity)
                ws.cell(row=row, column=5, value=record.production_date)
                ws.cell(row=row, column=6, value=record.shelf_life)
                
                if supplier:
                    ws.cell(row=row, column=7, value='')
                    ws.cell(row=row, column=8, value=supplier.name)
                    ws.cell(row=row, column=9, value=f"{supplier.address} {supplier.phone}")
                else:
                    ws.cell(row=row, column=8, value=record.supplier_name)
                    ws.cell(row=row, column=9, value=f"{record.supplier_address} {record.supplier_phone}")
                
                if record.certificate_no:
                    ws.cell(row=row, column=10, value='√')
                    ws.cell(row=row, column=11, value='√')
                    ws.cell(row=row, column=12, value='√')
                    ws.cell(row=row, column=13, value='√')
                
                if record.inspection_result and '合格' in record.inspection_result:
                    ws.cell(row=row, column=14, value='√')
                    ws.cell(row=row, column=15, value='√')
                
                ws.cell(row=row, column=16, value=record.inspector)
            
            wb.save(file_path)
            wb.close()
            return True

        except Exception as e:
            QMessageBox.critical(parent_widget, "导出失败", f"导出时发生错误:\n{str(e)}")
            return False

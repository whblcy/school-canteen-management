"""
V2 full-functional integration test.
Mirrors app.build_services() wiring, exercises the entire feature set against an
in-memory SQLite DB + real Excel report templates. ASCII-only source.
"""
import sys, os, tempfile
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
from datetime import datetime

ROOT = r"C:\Users\lcy\work\school-canteen-management"
sys.path.insert(0, ROOT)

from peewee import SqliteDatabase
import school_canteen.data.database as dbmod
mem = SqliteDatabase(":memory:", pragmas={"foreign_keys": 1})
dbmod.db.initialize(mem)

from school_canteen.data.models import (
    User, Category, Supplier, Ingredient, StockIn, StockOut,
    InventoryCheck, OperationLog, InspectionRecord, Inspector, CategoryMapping)
dbmod.db.connect(reuse_if_open=True)
dbmod.db.create_tables(
    [User, Category, Supplier, Ingredient, StockIn, StockOut,
     InventoryCheck, OperationLog, InspectionRecord, Inspector,
     CategoryMapping], safe=True)
from school_canteen.data.models import seed_default_data
seed_default_data()

from school_canteen.data.repositories.user_repository import UserRepository
from school_canteen.data.repositories.ingredient_repository import IngredientRepository
from school_canteen.data.repositories.catalog_repository import (
    CategoryRepository, SupplierRepository, CategoryMappingRepository)
from school_canteen.data.repositories.stock_repository import (
    StockInRepository, StockOutRepository, InventoryCheckRepository)
from school_canteen.data.repositories.inspection_repository import (
    InspectionRecordRepository, InspectorRepository)
from school_canteen.data.repositories.report_repository import (
    LogRepository, ReportRepository)

from school_canteen.services.auth_service import AuthService
from school_canteen.services.ingredient_service import IngredientService
from school_canteen.services.stock_service import StockService
from school_canteen.services.inspection_service import (
    InspectionService, InspectorService)
from school_canteen.services.report_service import ReportService
from school_canteen.services.catalog_service import (
    CategoryService, SupplierService, CategoryMappingService)
from school_canteen.services.data_management_service import DataManagementService
from school_canteen.services.user_service import UserService
from school_canteen.services.utility_services import (
    LogService, ExcelExportService, ReportExportService)
from school_canteen.core.session import Session
from school_canteen.core.exceptions import (
    AuthenticationError, ValidationError, DuplicateError, AuthorizationError,
    InsufficientStockError, ExpiredIngredientError, NotFoundError)
from school_canteen.utils.report_generator import ReportGenerator

# ---- wiring (mirror app.build_services) ----
user_repo = UserRepository(); category_repo = CategoryRepository()
supplier_repo = SupplierRepository(); ingredient_repo = IngredientRepository()
stock_in_repo = StockInRepository(); stock_out_repo = StockOutRepository()
inventory_repo = InventoryCheckRepository()
inspection_repo = InspectionRecordRepository(); inspector_repo = InspectorRepository()
mapping_repo = CategoryMappingRepository(); log_repo = LogRepository()
report_repo = ReportRepository()

auth_service = AuthService(user_repo, log_repo)
category_service = CategoryService(category_repo, log_repo)
supplier_service = SupplierService(supplier_repo, log_repo)
mapping_service = CategoryMappingService(mapping_repo, category_repo, log_repo)
ingredient_service = IngredientService(
    ingredient_repo, category_repo, supplier_repo, log_repo)
stock_service = StockService(
    ingredient_repo, stock_in_repo, stock_out_repo, inventory_repo,
    category_repo, supplier_repo, log_repo)
inspection_service = InspectionService(
    inspection_repo, inspector_repo, stock_in_repo, log_repo)
inspector_service = InspectorService(inspector_repo, log_repo)
report_service = ReportService(
    ingredient_repo, stock_in_repo, stock_out_repo, report_repo)
data_mgmt_service = DataManagementService(report_repo, log_repo)
user_service = UserService(user_repo, log_repo)
log_service = LogService(log_repo)
excel_service = ExcelExportService(ingredient_service, stock_service)
report_export_service = ReportExportService(
    stock_in_repo, stock_out_repo, ingredient_repo,
    inventory_repo, inspection_repo)

passed = []
failed = []

def check(name, cond, extra=""):
    if cond:
        passed.append(name); print("PASS:", name)
    else:
        failed.append(name); print("FAIL:", name, extra)

def expect_raise(name, exc_type, fn):
    try:
        fn()
    except exc_type:
        passed.append(name); print("PASS:", name)
    except Exception as e:
        failed.append(name); print("FAIL:", name, "wrong exc:", repr(e))
    else:
        failed.append(name); print("FAIL:", name, "no exception raised")

# ============ A. seed ============
check("seed admin exists", user_repo.get_by_username("admin") is not None)
check("seed categories >=8", len(category_service.get_all()) >= 8)
check("seed inspectors >=4", len(inspector_service.get_all()) >= 4)

# ============ B. auth ============
admin = auth_service.login("admin", "admin123")
check("login returns user", admin is not None)
check("session is admin", Session.current_user is not None and Session.role == "admin")
check("admin can_manage_users", Session.can_manage_users is True)
expect_raise("login wrong pw raises", AuthenticationError,
             lambda: auth_service.login("admin", "wrongpw"))
expect_raise("login unknown user raises", AuthenticationError,
             lambda: auth_service.login("nobody", "x"))
check("change_password ok", auth_service.change_password(admin.id, "admin123", "newadm1") is True)
check("login new pw works", auth_service.login("admin", "newadm1") is not None)
check("change_password back", auth_service.change_password(admin.id, "newadm1", "admin123") is True)

# ============ C. user management ============
mgr = user_service.create_user("manager1", "manager1", "库管", "manager")
norm = user_service.create_user("normuser", "norm1234", "普通员", "user")
check("create manager role", mgr.role == "manager")
check("create normal role", norm.role == "user")
expect_raise("duplicate user raises", DuplicateError,
             lambda: user_service.create_user("normuser", "norm1234", "x", "user"))
expect_raise("short pw raises", ValidationError,
             lambda: user_service.create_user("short", "123", "x", "user"))
check("update_role ok", user_service.update_role(norm.id, "manager") is True)
expect_raise("cannot demote admin", ValidationError,
             lambda: user_service.update_role(admin.id, "user"))
check("disable user ok", user_service.update_status(norm.id, 0) is True)
check("disabled user cannot auth", user_repo.authenticate("normuser", "norm1234") is None)
check("re-enable user", user_service.update_status(norm.id, 1) is True)
check("reset pw ok", user_service.reset_password(norm.id, "reset12") is True)
check("login after reset pw", auth_service.login("normuser", "reset12") is not None)
# authorization gate: normal user cannot manage users
Session.current_user = norm
expect_raise("normal user blocked from user mgmt", AuthorizationError,
             lambda: user_service.create_user("hacker", "hacker1", "x", "user"))
Session.current_user = admin
expect_raise("cannot delete admin", ValidationError,
             lambda: user_service.delete_user(admin.id))

# ============ D. category / supplier / mapping ============
catx = category_service.create("测试分类X")
check("create category", catx is not None)
expect_raise("duplicate category raises", DuplicateError,
             lambda: category_service.create("测试分类X"))
supx = supplier_service.create("测试供应商X", "张三", "13800000000")
check("create supplier", supx is not None)
meat = category_repo.get_by_name("肉类")
check("seed has 肉类", meat is not None)
mp = mapping_service.create("自定义来源X", meat.id)
check("create mapping", mp is not None)

# ============ E. ingredient ============
ingA = ingredient_service.create_ingredient(
    "测试食材A", "蔬菜类", "公斤", safety_stock=10, supplier_name="测试供应商X")
check("ingredient auto category", ingA.category is not None)
check("ingredient auto supplier", ingA.supplier is not None)
expect_raise("duplicate ingredient raises", DuplicateError,
             lambda: ingredient_service.create_ingredient("测试食材A", "蔬菜类", "公斤"))
low = ingredient_service.get_low_stock()
check("low stock detected", any(i.id == ingA.id for i in low))
check("update ingredient ok", ingredient_service.update_ingredient(ingA.id, name="测试食材A改") is True)

# ============ F. stock core + rollback (admin) ============
ingS = ingredient_service.create_ingredient("回冲测试食材", "蔬菜类", "个")
check("ingS start stock 0", ingredient_repo.get_stock(ingS.id) == 0)
rin = stock_service.stock_in(ingS.id, 50, 2.0)
check("after in stock 50", ingredient_repo.get_stock(ingS.id) == 50)
rout = stock_service.stock_out(ingS.id, 20)
check("after out stock 30", ingredient_repo.get_stock(ingS.id) == 30)
check("out weighted price", abs(rout.unit_price - 2.0) < 1e-9)
stock_service.delete_stock_out(rout.id)
check("after del out stock 50", ingredient_repo.get_stock(ingS.id) == 50)
stock_service.delete_stock_in(rin.id)
check("after del in stock 0", ingredient_repo.get_stock(ingS.id) == 0)
check("rin gone", StockIn.get_or_none(StockIn.id == rin.id) is None)
check("rout gone", StockOut.get_or_none(StockOut.id == rout.id) is None)

# ============ G. permission gating + expiry + insufficient ============
# admin custom time honored
rec_in = stock_service.stock_in(ingA.id, 100, 5.0, created_at="2026-06-10 08:00:00")
check("admin custom time honored", rec_in.created_at == datetime(2026, 6, 10, 8, 0, 0))
rec_out = stock_service.stock_out(ingA.id, 30, created_at="2026-06-10 14:00:00")
check("ingA stock 70", ingredient_repo.get_stock(ingA.id) == 70)
# normal user custom time ignored (writes now)
Session.current_user = norm
rec_norm = stock_service.stock_in(ingA.id, 10, 1.0, created_at="2020-01-01 00:00:00")
check("normal time ignored", rec_norm.created_at.year == datetime.now().year)
check("ingA stock 80", ingredient_repo.get_stock(ingA.id) == 80)
Session.current_user = admin
# expired ingredient: normal blocked, manager bypasses
ingB = ingredient_service.create_ingredient("过期食材B", "肉类", "个")
stock_service.stock_in(ingB.id, 50, 1.0, expiry_date="2020-01-01")
Session.current_user = norm
expect_raise("normal blocked expired out", ExpiredIngredientError,
             lambda: stock_service.stock_out(ingB.id, 10))
Session.current_user = mgr
rout_b = stock_service.stock_out(ingB.id, 10)  # manager bypass
check("manager bypass expiry", rout_b is not None)
Session.current_user = admin
expect_raise("insufficient stock raises", InsufficientStockError,
             lambda: stock_service.stock_out(ingA.id, 999999))

# ============ H. inventory check ============
chk = stock_service.inventory_check(ingA.id, actual_stock=50)
check("inventory check record", chk is not None)
check("stock set to actual", ingredient_repo.get_stock(ingA.id) == 50)
check("difference computed", chk.difference == (50 - 80))  # 80 was before check

# ============ I. inspection upsert + query ============
insp_stock = stock_service.stock_in(ingA.id, 5, 3.0)
insp = inspection_service.upsert_record(
    stock_in_id=insp_stock.id, ingredient_id=ingA.id, quantity=5, unit="公斤",
    inspection_result="合格", inspector="张三", inspection_date="2026-06-10")
check("inspection upsert", insp is not None)
check("inspection linked", insp.stock_in_id == insp_stock.id)
# also write inspection for rec_in (June) so June report non-empty
inspection_service.upsert_record(
    stock_in_id=rec_in.id, ingredient_id=ingA.id, quantity=100, unit="公斤",
    inspection_result="合格", inspector="张三", inspection_date="2026-06-10")
rng = inspection_service.get_by_date_range("2026-06-01", "2026-06-30")
check("inspection date range", any(r.stock_in_id == rec_in.id for r in rng))

# ============ J. Excel inspection import (full chain) ============
import openpyxl
tmpx = os.path.join(tempfile.mkdtemp(), "imp.xlsx")
wb = openpyxl.Workbook(); ws = wb.active
ws.append(["产品名称", "单位", "数量", "生产日期", "供货单位", "记录人"])
ws.append([]); ws.append([])
ws.append(["进口牛肉", "公斤", 20, "2026-06-01", "海外供应商", "李四"])
wb.save(tmpx)
from school_canteen.utils.excel_handler import ExcelImporter
ok, msg = ExcelImporter.import_inspection_records(
    None, stock_service, inspection_service, file_path=tmpx)
check("excel import returns ok", ok is True)
imp_ing = ingredient_repo.get_by_name("进口牛肉")
check("import auto ingredient", imp_ing is not None)
imp_si = StockIn.select().where(StockIn.ingredient == imp_ing.id).first() if imp_ing else None
check("import created stock_in", imp_si is not None and imp_si.quantity == 20)
imp_insp = inspection_repo.get_by_stock_in_id(imp_si.id) if imp_si else None
check("import wrote inspection record", imp_insp is not None and imp_insp.inspector == "李四")

# ============ K. reports ============
tmp = tempfile.mkdtemp()
import os as _os
try:
    out_daily = _os.path.join(tmp, "daily.xlsx")
    report_export_service.export_daily_stock_sheet(out_daily, 2026, 6, 10)
    _wb = openpyxl.load_workbook(out_daily); _ws = _wb["出入库表"]
    check("daily D2 date", _ws["D2"].value == datetime(2026, 6, 10))
    check("daily in name", _ws.cell(row=5, column=2).value == "测试食材A改")
    check("daily in qty", _ws.cell(row=5, column=3).value == 100)
    check("daily out qty", _ws.cell(row=5, column=12).value == 30)

    out_month = _os.path.join(tmp, "month.xlsx")
    report_export_service.export_monthly_summary(out_month, 2026, 6)
    _wb = openpyxl.load_workbook(out_month)
    _win = _wb["2.每月采购食材汇总表"]
    _wout = _wb["4.每月发出食材汇总表"]
    check("monthly in qty", _win.cell(row=4, column=4).value == 100)
    check("monthly in amount", _win.cell(row=4, column=6).value == 500)
    check("monthly out qty", _wout.cell(row=4, column=4).value == 30)
    check("monthly out amount", _wout.cell(row=4, column=6).value == 150)

    out_fin = _os.path.join(tmp, "fin.xlsx")
    report_export_service.export_financial_report(out_fin, 2026)
    _wb = openpyxl.load_workbook(out_fin)
    _wj = None
    for _n in _wb.sheetnames:
        if "6月" in _n.replace(" ", ""):
            _wj = _wb[_n]; break
    check("financial june sheet", _wj is not None)
    if _wj is not None:
        check("financial B5", _wj["B5"].value == 500)
        check("financial E5", _wj["E5"].value == 500)

    out_insp = _os.path.join(tmp, "insp.xlsx")
    report_export_service.export_inspection_report(out_insp, 2026, 6)
    _wb = openpyxl.load_workbook(out_insp); _ws = _wb["Sheet1"]
    check("insp report name", _ws.cell(row=4, column=2).value == "测试食材A改")
    check("insp report result", _ws.cell(row=4, column=14).value == "合格")
except Exception as e:
    failed.append("reports"); print("FAIL: reports ->", repr(e))

# ============ L. overview stats ============
ov = report_service.get_overview_stats()
check("overview has total_ingredients", "total_ingredients" in ov and isinstance(ov["total_ingredients"], int))
check("overview has inventory_value", "inventory_value" in ov)

# ============ M. data management ============
expect_raise("clear wrong phrase raises", ValidationError,
             lambda: data_mgmt_service.clear_data("stock_records", "错误的"))
data_mgmt_service.clear_data("stock_records", "确认删除")
check("stock cleared", StockIn.select().count() == 0 and StockOut.select().count() == 0)
check("inspection cleared", InspectionRecord.select().count() == 0)
check("stock reset to 0", ingredient_repo.get_stock(ingA.id) == 0)
check("users survive clear", user_repo.get_by_username("admin") is not None)
data_mgmt_service.clear_data("all_data", "确认删除")
check("all_data reseeds admin", user_repo.get_by_username("admin") is not None)
check("all_data clears stock", StockIn.select().count() == 0)

# ============ N. audit log ============
logs = log_service.get_all()
login_log = [l for l in logs if l.action == "登录"]
check("login log exists", len(login_log) > 0)
check("login log has user", any(l.user is not None for l in login_log))

print()
print("=== RESULT: %d passed, %d failed ===" % (len(passed), len(failed)))
if failed:
    print("FAILED:", failed)
    sys.exit(1)
print("ALL FUNCTIONAL CHECKS PASSED")

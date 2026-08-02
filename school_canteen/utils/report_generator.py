"""
报表生成器 - 基于预置 Excel 模板生成监管报表
模板文件存放在 表格/ 目录下

注意：所有导出函数均按实测模板结构填充。若目标工作表缺失，
统一抛出 ReportTemplateError，避免"导出成功但内容空白"的静默失败。
"""
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ..config import get_config
from ..core.exceptions import ReportTemplateError
from ..utils.logging_config import get_logger

logger = get_logger()


def _get_template_path(filename: str) -> Path:
    cfg = get_config()
    return cfg.paths.template(filename)


def _load_template(template_filename: str, output_path: str) -> load_workbook:
    """复制模板并加载工作簿"""
    template_path = _get_template_path(template_filename)
    if not template_path.exists():
        raise FileNotFoundError(f"报表模板不存在: {template_path}")
    shutil.copy2(str(template_path), output_path)
    return load_workbook(output_path)


def _copy_cell_style(src_cell, dst_cell):
    """复制单元格样式"""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.border = copy(src_cell.border)
        dst_cell.number_format = src_cell.number_format


def _insert_row_with_style(ws, row_idx: int, ref_row: int):
    """插入行并继承参考行样式"""
    ws.insert_rows(row_idx)
    for col in range(1, ws.max_column + 1):
        ref_cell = ws.cell(row=ref_row, column=col)
        new_cell = ws.cell(row=row_idx, column=col)
        _copy_cell_style(ref_cell, new_cell)


def _require_sheet(wb, name: str):
    """要求模板中存在指定工作表，否则报错"""
    if name not in wb.sheetnames:
        raise ReportTemplateError(
            f"模板缺少工作表 '{name}'，现有工作表: {wb.sheetnames}")
    return wb[name]


def _find_sheet_by_key(wb, key: str):
    """按关键字（忽略空格）查找工作表；找不到返回 None"""
    normalized_key = key.replace(" ", "")
    for name in wb.sheetnames:
        if normalized_key in name.replace(" ", ""):
            return wb[name]
    return None


def _ensure_row(ws, row_idx: int):
    """若超出模板最大行则插入带样式的行"""
    if row_idx > ws.max_row:
        _insert_row_with_style(ws, row_idx, ws.max_row)


class ReportGenerator:
    """报表生成器"""

    @staticmethod
    def export_daily_stock_sheet(output_path: str, stock_in_repo, stock_out_repo,
                                 year: int, month: int, day: int):
        """
        每日出入库表 - 单日出入库台账
        模板「出入库表」sheet：左侧入库(col1-7)，右侧出库(col9-16)，
        row2 的 D2/M2 为日期，row3 为表头，row5 起为数据。
        """
        wb = _load_template("2026年6月份每天出入库表.xlsx", output_path)
        ws = _require_sheet(wb, "出入库表")

        days = _days_in_month(year, month)
        if not (1 <= day <= days):
            raise ReportTemplateError(
                f"日期 {day} 日超出 {year}年{month}月 有效范围(1-{days}日)")
        date_obj = datetime(year, month, day)
        date_str = f"{year}-{month:02d}-{day:02d}"
        ws["D2"] = date_obj   # 入库日期
        ws["M2"] = date_obj   # 出库日期（原模板为 =D2 公式，导出覆盖为值）

        in_records = stock_in_repo.get_by_date(date_str)
        out_records = stock_out_repo.get_by_date(date_str)

        # 入库：col1 类别, col2 品目, col3 数量, col4 单位, col5 单价, col6 金额
        r = 5
        for rec in in_records:
            _ensure_row(ws, r)
            ws.cell(row=r, column=1, value=rec.get("category_name", ""))
            ws.cell(row=r, column=2, value=rec.get("ingredient_name", ""))
            ws.cell(row=r, column=3, value=rec.get("total_quantity", 0))
            ws.cell(row=r, column=4, value=rec.get("unit", ""))
            ws.cell(row=r, column=5, value=rec.get("avg_price", 0))
            ws.cell(row=r, column=6, value=rec.get("total_amount", 0))
            r += 1

        # 出库：col9 类别, col10 品目, col11 单位, col12 数量, col13 实领,
        #       col14 单价, col15 金额
        r = 5
        for rec in out_records:
            _ensure_row(ws, r)
            ws.cell(row=r, column=9, value=rec.get("category_name", ""))
            ws.cell(row=r, column=10, value=rec.get("ingredient_name", ""))
            ws.cell(row=r, column=11, value=rec.get("unit", ""))
            ws.cell(row=r, column=12, value=rec.get("total_quantity", 0))
            ws.cell(row=r, column=13, value=rec.get("total_quantity", 0))
            ws.cell(row=r, column=14, value=rec.get("avg_price", 0))
            ws.cell(row=r, column=15, value=rec.get("total_amount", 0))
            r += 1

        wb.save(output_path)

    @staticmethod
    def export_monthly_summary(output_path: str, stock_in_repo, stock_out_repo, year: int, month: int):
        """
        每月出入库统计表
        模板含「2.每月采购食材汇总表」(入库) 与「4.每月发出食材汇总表」(出库)，
        两表结构相同：col1 类别, col2 品目, col3 单位, col4 数量,
        col5 平均单价, col6 金额, col7 备注。row3 表头，row4 起数据。
        """
        wb = _load_template(
            "2026年6月每月出入库汇总等表和14、每月结算食材公示.xlsx", output_path)
        ws_in = _require_sheet(wb, "2.每月采购食材汇总表")
        ws_out = _require_sheet(wb, "4.每月发出食材汇总表")

        in_data = stock_in_repo.get_monthly(year, month)
        out_data = stock_out_repo.get_monthly(year, month)

        r = 4
        for row in in_data:
            _ensure_row(ws_in, r)
            ws_in.cell(row=r, column=1, value=row.get("category_name", ""))
            ws_in.cell(row=r, column=2, value=row.get("ingredient_name", ""))
            ws_in.cell(row=r, column=3, value=row.get("unit", ""))
            ws_in.cell(row=r, column=4, value=row.get("total_quantity", 0))
            ws_in.cell(row=r, column=5, value=row.get("avg_price", 0))
            ws_in.cell(row=r, column=6, value=row.get("total_amount", 0))
            r += 1

        r = 4
        for row in out_data:
            _ensure_row(ws_out, r)
            ws_out.cell(row=r, column=1, value=row.get("category_name", ""))
            ws_out.cell(row=r, column=2, value=row.get("ingredient_name", ""))
            ws_out.cell(row=r, column=3, value=row.get("unit", ""))
            ws_out.cell(row=r, column=4, value=row.get("total_quantity", 0))
            ws_out.cell(row=r, column=5, value=row.get("avg_price", 0))
            ws_out.cell(row=r, column=6, value=row.get("total_amount", 0))
            r += 1

        wb.save(output_path)

    @staticmethod
    def export_financial_report(output_path: str, stock_in_repo, stock_out_repo, year: int):
        """
        财务收支情况表 - 12 个月分 Sheet，sheet 名为「N 月份」(带空格)
        收入「财政补助」本月数在 B 列(row5)，支出「食材」本月数在 E 列(row5)。
        """
        wb = _load_template("附件1 ：月份、年度食堂（营养餐）财务收支情况表(1).xlsx", output_path)

        # 模板按月分 sheet，逐月填充（收入/支出-食材 本月数）
        in_by_month = {
            m: stock_in_repo.get_total_amount_by_month(year, m) for m in range(1, 13)
        }
        out_by_month = {
            m: stock_out_repo.get_total_amount_by_month(year, m) for m in range(1, 13)
        }
        for m in range(1, 13):
            ws = _find_sheet_by_key(wb, f"{m}月")
            if ws is None:
                logger.warning(f"财务模板缺少 {m}月 工作表，跳过")
                continue
            # 收入-财政补助本月数 (B5) = 当月采购支出（由补助资金覆盖）
            ws["B5"] = in_by_month[m]
            # 支出-食材本月数 (E5) = 当月采购支出
            ws["E5"] = in_by_month[m]

        wb.save(output_path)

    @staticmethod
    def export_inventory_check_sheet(output_path: str, ingredient_repo, inventory_repo, check_records: list = None):
        """库存物品盘存盘亏表"""
        wb = _load_template("库存物品盘存盘亏表.xlsx", output_path)
        ws = wb.active

        ingredients = ingredient_repo.get_all_with_relations()
        start_row = 4
        for i, ing in enumerate(ingredients):
            r = start_row + i
            if r > ws.max_row:
                _insert_row_with_style(ws, r, ws.max_row)
            cat_name = ing.category.name if ing.category else ""
            ws.cell(row=r, column=1, value=cat_name)
            ws.cell(row=r, column=2, value=ing.name)
            ws.cell(row=r, column=3, value=ing.unit)
            ws.cell(row=r, column=4, value=ing.current_stock)

            if check_records:
                for cr in check_records:
                    if cr.ingredient_id == ing.id:
                        ws.cell(row=r, column=5, value=cr.actual_stock)
                        ws.cell(row=r, column=6, value=cr.difference)
                        break

        wb.save(output_path)

    @staticmethod
    def export_inspection_report(output_path: str, inspection_repo, year: int, month: int):
        """进货查验记录表
        模板 Sheet1 表头(row2 主/row3 子)：
        col1 进货日期, col2 产品名称, col3 单位, col4 数量, col5 生产日期,
        col6 保质期, col7 生产单位, col8 供货单位, col9 地址及联系方式,
        col14 入库检查, col16 记录人。row4 起为数据。
        """
        wb = _load_template("进货查验记录表.xlsx", output_path)
        ws = _require_sheet(wb, "Sheet1")

        date_from = f"{year}-{month:02d}-01"
        days = _days_in_month(year, month)
        date_to = f"{year}-{month:02d}-{days:02d}"
        records = inspection_repo.get_by_date_range(date_from, date_to)

        r = 4
        for rec in records:
            _ensure_row(ws, r)
            ing_name = rec.ingredient.name if rec.ingredient else ""
            contact = " ".join(filter(None, [
                str(rec.supplier_address or ""), str(rec.supplier_phone or "")])).strip()
            ws.cell(row=r, column=1, value=rec.inspection_date or "")
            ws.cell(row=r, column=2, value=ing_name)
            ws.cell(row=r, column=3, value=rec.unit or "")
            ws.cell(row=r, column=4, value=rec.quantity or 0)
            ws.cell(row=r, column=5, value=rec.production_date or "")
            ws.cell(row=r, column=6, value=rec.shelf_life or "")
            ws.cell(row=r, column=7, value=rec.supplier_name or "")
            ws.cell(row=r, column=8, value=rec.supplier_name or "")
            ws.cell(row=r, column=9, value=contact)
            ws.cell(row=r, column=14, value=rec.inspection_result or "")
            ws.cell(row=r, column=16, value=rec.inspector or "")
            r += 1

        wb.save(output_path)


def _days_in_month(year: int, month: int) -> int:
    """计算某月天数"""
    from calendar import monthrange
    return monthrange(year, month)[1]

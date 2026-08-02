"""
Excel 导入导出工具 - 适配新数据层
通过 Service 层访问数据，保持业务规则一致性
"""
import os
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import QFileDialog, QMessageBox, QProgressDialog, QDialog, QVBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QPushButton, QHBoxLayout, QHeaderView
from PyQt6.QtCore import Qt

from ..config import get_config

# macOS 风格配色
MAC_BLUE = "0071e3"
MAC_GREEN = "34c759"
MAC_RED = "ff3b30"
MAC_ORANGE = "ff9500"
MAC_GRAY = "86868b"
MAC_LIGHT_GRAY = "f5f5f7"
MAC_BORDER = "d2d2d7"


def _header_style():
    return {
        "fill": PatternFill(start_color=MAC_LIGHT_GRAY, end_color=MAC_LIGHT_GRAY, fill_type="solid"),
        "font": Font(color="1d1d1f", bold=True, size=12, name="-apple-system"),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border": Border(bottom=Side(style="thin", color=MAC_BORDER)),
    }


def _cell_style():
    return {
        "font": Font(color="1d1d1f", size=11, name="-apple-system"),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border": Border(bottom=Side(style="thin", color="f0f0f0")),
    }


class ExcelExporter:
    """Excel 导出器"""

    @staticmethod
    def export_ingredients(parent, ingredients: list, file_path: str = None) -> bool:
        """导出食材余量信息"""
        if not file_path:
            file_path, _ = QFileDialog.getSaveFileName(
                parent, "导出食材余量信息",
                f"食材余量_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)",
            )
        if not file_path:
            return False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "食材余量"
            hs = _header_style()
            cs = _cell_style()

            headers = ["ID", "食材名称", "分类", "规格", "单位", "当前库存", "安全库存", "库存状态", "供应商"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = hs["fill"]
                cell.font = hs["font"]
                cell.alignment = hs["alignment"]
                cell.border = hs["border"]

            for row, ing in enumerate(ingredients, 2):
                status = "正常" if ing.current_stock > ing.safety_stock else "库存不足"
                cat_name = ing.category.name if ing.category else ""
                sup_name = ing.supplier.name if ing.supplier else ""
                data = [ing.id, ing.name, cat_name, ing.specification,
                        ing.unit, ing.current_stock, ing.safety_stock, status, sup_name]
                for col, val in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.font = cs["font"]
                    cell.alignment = cs["alignment"]
                    cell.border = cs["border"]
                if ing.current_stock <= ing.safety_stock:
                    for col in range(1, 10):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="fff5f5", end_color="fff5f5", fill_type="solid")
                        ws.cell(row=row, column=col).font = Font(color=MAC_RED, size=11, name="-apple-system")

            for i, w in enumerate([8, 22, 14, 16, 10, 14, 14, 14, 18], 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            summary_row = len(ingredients) + 3
            ws.cell(row=summary_row, column=1, value="汇总信息").font = Font(bold=True, size=13, color=MAC_BLUE, name="-apple-system")
            ws.cell(row=summary_row + 1, column=1, value="食材种类总数:")
            ws.cell(row=summary_row + 1, column=2, value=len(ingredients))
            low_count = sum(1 for ing in ingredients if ing.current_stock <= ing.safety_stock)
            ws.cell(row=summary_row + 2, column=1, value="库存预警数量:")
            ws.cell(row=summary_row + 2, column=2, value=low_count).font = Font(color=MAC_RED, bold=True, name="-apple-system")

            wb.save(file_path)
            return True
        except Exception as e:
            QMessageBox.critical(parent, "导出失败", f"导出时发生错误:\n{str(e)}")
            return False

    @staticmethod
    def export_stock_records(parent, records: list, record_type: str = "in", file_path: str = None) -> bool:
        """导出入库/出库记录"""
        if not file_path:
            file_path, _ = QFileDialog.getSaveFileName(
                parent, f"导出{'入库' if record_type == 'in' else '出库'}记录",
                f"{'入库' if record_type == 'in' else '出库'}记录_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)",
            )
        if not file_path:
            return False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "记录"
            hs = _header_style()
            cs = _cell_style()

            if record_type == "in":
                headers = ["ID", "食材", "数量", "单价", "总价", "供应商", "批次号", "生产日期", "保质期", "操作人", "备注", "时间"]
            else:
                headers = ["ID", "食材", "数量", "单价", "总价", "用途", "领用部门", "操作人", "备注", "时间"]

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = hs["fill"]
                cell.font = hs["font"]
                cell.border = hs["border"]

            for row, r in enumerate(records, 2):
                ing_name = r.ingredient.name if r.ingredient else ""
                if record_type == "in":
                    sup_id = r.supplier_id if r.supplier else ""
                    data = [r.id, ing_name, r.quantity, r.unit_price, r.total_price,
                            sup_id, r.batch_number, r.production_date, r.expiry_date,
                            r.operator, r.remark, r.created_at]
                else:
                    data = [r.id, ing_name, r.quantity, r.unit_price, r.total_price,
                            r.purpose, r.department, r.operator, r.remark, r.created_at]
                for col, val in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.font = cs["font"]
                    cell.border = cs["border"]

            wb.save(file_path)
            return True
        except Exception as e:
            QMessageBox.critical(parent, "导出失败", str(e))
            return False

    @staticmethod
    def create_template(parent) -> bool:
        """创建导入模板"""
        file_path, _ = QFileDialog.getSaveFileName(
            parent, "保存导入模板", "食材导入模板.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "食材导入模板"

            ws.merge_cells("A1:F1")
            ws["A1"].value = "学校食堂食材导入模板"
            ws["A1"].font = Font(size=18, bold=True, color=MAC_BLUE, name="-apple-system")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 36

            ws.merge_cells("A2:F2")
            ws["A2"].value = "请按照下方格式填写食材信息，带 * 号为必填项"
            ws["A2"].font = Font(size=11, color=MAC_GRAY, name="-apple-system")
            ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 24

            headers = [
                ("食材名称*", "食材的完整名称"), ("分类", "所属分类"),
                ("规格", "规格描述"), ("单位*", "计量单位"),
                ("安全库存", "库存预警阈值"), ("供应商", "供应商名称"),
            ]
            h_fill = PatternFill(start_color=MAC_LIGHT_GRAY, end_color=MAC_LIGHT_GRAY, fill_type="solid")
            h_font = Font(color="1d1d1f", bold=True, size=12, name="-apple-system")
            h_border = Border(bottom=Side(style="medium", color=MAC_BORDER))
            for col, (h, _) in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=h)
                cell.fill = h_fill
                cell.font = h_font
                cell.border = h_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[4].height = 28

            examples = [
                ["大白菜", "蔬菜类", "新鲜", "斤", 50, "绿源蔬菜批发"],
                ["五花肉", "肉类", "精品", "斤", 30, "鸿运肉业"],
                ["鸡蛋", "蛋类", "散养土鸡蛋", "个", 200, "阳光养殖场"],
                ["东北大米", "粮油类", "五常稻花香", "袋", 20, "金穗粮油"],
            ]
            for row_idx, ex in enumerate(examples, 5):
                for col_idx, val in enumerate(ex, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = Font(size=11, name="-apple-system")
                    cell.border = Border(bottom=Side(style="thin", color="f0f0f0"))

            for i, w in enumerate([18, 14, 18, 10, 14, 18], 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            wb.save(file_path)
            return True
        except Exception as e:
            QMessageBox.critical(parent, "创建模板失败", str(e))
            return False


class ExcelImporter:
    """Excel 导入器 - 通过 Service 层写入数据"""

    @staticmethod
    def import_ingredients(parent, ingredient_service, file_path: str = None):
        """从 Excel 导入食材信息"""
        if not file_path:
            file_path, _ = QFileDialog.getOpenFileName(parent, "导入食材信息", "", "Excel Files (*.xlsx)")
        if not file_path:
            return False, "未选择文件"

        try:
            wb = load_workbook(file_path)
            ws = wb.active
            expected = ["食材名称", "分类", "规格", "单位", "安全库存", "供应商"]
            actual = [cell.value for cell in ws[1]]
            header_map = {h: i for i, h in enumerate(actual) if h in expected}

            if "食材名称" not in header_map or "单位" not in header_map:
                return False, "Excel 格式不正确，必须包含'食材名称'和'单位'列"

            success, skip, error = 0, 0, 0
            errors = []
            existing = {ing.name.lower() for ing in ingredient_service.get_all_ingredients()}
            imported = set()

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    name = row[header_map.get("食材名称", 0)]
                    if not name:
                        continue
                    name_str = str(name).strip()

                    if name_str.lower() in existing:
                        skip += 1
                        errors.append(f"第{row_idx}行: 食材'{name_str}'已存在，跳过")
                        continue
                    if name_str.lower() in imported:
                        skip += 1
                        errors.append(f"第{row_idx}行: 食材'{name_str}'文件内重复，跳过")
                        continue
                    imported.add(name_str.lower())

                    category_name = row[header_map.get("分类", 1)] if "分类" in header_map else None
                    specification = row[header_map.get("规格", 2)] if "规格" in header_map else ""
                    unit = row[header_map.get("单位", 3)] if "单位" in header_map else ""
                    safety_stock = 0
                    if "安全库存" in header_map:
                        try:
                            safety_stock = float(row[header_map["安全库存"]])
                        except (ValueError, TypeError):
                            safety_stock = 0
                    supplier_name = row[header_map.get("供应商", 5)] if "供应商" in header_map else None

                    ingredient_service.create_ingredient(
                        name=name_str, category_name=category_name,
                        unit=str(unit) if unit else "个",
                        specification=str(specification) if specification else "",
                        safety_stock=safety_stock,
                        supplier_name=supplier_name,
                    )
                    existing.add(name_str.lower())
                    success += 1
                except Exception as e:
                    error += 1
                    errors.append(f"第{row_idx}行: {str(e)}")

            msg = f"导入完成!\n成功: {success} 条\n跳过(重复): {skip} 条\n失败: {error} 条"
            if errors:
                msg += "\n\n详情:\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    msg += f"\n... 还有 {len(errors) - 5} 条"
            return True, msg
        except Exception as e:
            return False, f"导入失败: {str(e)}"

    @staticmethod
    def import_sales_orders(parent, stock_service, file_path: str = None):
        """从销售订单 Excel 导入入库数据"""
        if not file_path:
            file_path, _ = QFileDialog.getOpenFileName(parent, "导入销售订单", "", "Excel Files (*.xlsx)")
        if not file_path:
            return False, "未选择文件"

        try:
            progress = QProgressDialog("正在解析Excel文件...", "取消", 0, 100, parent)
            progress.setWindowTitle("导入销售订单")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(10)

            wb = load_workbook(file_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            header_map = {}
            for i, h in enumerate(headers):
                header_map[str(h).strip() if h else ""] = i

            required = ["商品名称", "数量", "单位", "含税单价", "价税合计"]
            missing = [h for h in required if h not in header_map]
            if missing:
                return False, f"缺少必需列: {', '.join(missing)}"

            progress.setValue(15)
            progress.setLabelText("正在处理合并单元格...")

            all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False))
            merged_fill = {}
            for mr in ws.merged_cells.ranges:
                min_col, min_row = mr.min_col - 1, mr.min_row - 1
                max_col, max_row = mr.max_col - 1, mr.max_row - 1
                top_val = all_rows[min_row][min_col].value
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        merged_fill[(r, c)] = top_val

            row_values = []
            for row_idx, row in enumerate(all_rows):
                vals = []
                for col_idx, cell in enumerate(row):
                    vals.append(merged_fill.get((row_idx, col_idx), cell.value))
                row_values.append(vals)

            progress.setValue(20)
            progress.setLabelText("正在解析数据...")

            parsed_data = []
            total_rows = len(row_values) - 1
            progress.setMaximum(total_rows + 20)

            def parse_date(val):
                if not val:
                    return "", ""
                if hasattr(val, "strftime"):
                    return val.strftime("%Y-%m-%d"), val.strftime("%Y-%m-%d 08:00:00")
                val_str = str(val).strip()
                if not val_str:
                    return "", ""
                import re
                for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y",
                            "%Y.%m.%d", "%Y年%m月%d日", "%Y-%m-%d %H:%M:%S",
                            "%Y/%m/%d %H:%M:%S", "%m-%d-%Y", "%d-%m-%Y"]:
                    try:
                        d = datetime.strptime(val_str, fmt)
                        return d.strftime("%Y-%m-%d"), d.strftime("%Y-%m-%d 08:00:00")
                    except Exception:
                        continue
                m = re.match(r"(\d{4})[/.年](\d{1,2})[/.月](\d{1,2})", val_str)
                if m:
                    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
                    return f"{y:04d}-{mo:02d}-{d:02d}", f"{y:04d}-{mo:02d}-{d:02d} 08:00:00"
                return "", ""

            for i in range(1, len(row_values)):
                if progress.wasCanceled():
                    return False, "用户取消导入"
                progress.setValue(i)
                row = row_values[i]
                try:
                    product_name = row[header_map.get("商品名称", 0)]
                    if not product_name:
                        continue
                    quantity = row[header_map.get("数量", 0)]
                    unit = row[header_map.get("单位", 0)]
                    unit_price = row[header_map.get("含税单价", 0)]
                    total_price = row[header_map.get("价税合计", 0)]
                    if quantity is None or unit is None or unit_price is None:
                        continue
                    quantity = float(quantity)
                    unit_price = float(unit_price)
                    total_price = float(total_price) if total_price is not None else quantity * unit_price

                    category_name = row[header_map.get("商品类别", 0)]
                    supplier_name = row[header_map.get("生产单位（空白部分自己查验收单）", 0)] or "未知供应商"
                    delivery_date = row[header_map.get("发货日期", 0)]
                    stockin_date, stockin_created_at = parse_date(delivery_date)
                    if not stockin_date:
                        stockin_date = datetime.now().strftime("%Y-%m-%d")
                        stockin_created_at = datetime.now().strftime("%Y-%m-%d 08:00:00")

                    batch_number = row[header_map.get("单据编号", 0)] or ""
                    production_date = row[header_map.get("生鲜日期", 0)] or ""
                    expiry_days = row[header_map.get("保质期(天)", 0)]
                    operator = row[header_map.get("业务员", 0)] or ""

                    expiry_date = ""
                    if production_date and expiry_days:
                        try:
                            prod_str, _ = parse_date(production_date)
                            if prod_str:
                                prod_d = datetime.strptime(prod_str, "%Y-%m-%d")
                                expiry_date = (prod_d + timedelta(days=int(float(expiry_days)))).strftime("%Y-%m-%d")
                        except Exception:
                            pass

                    parsed_data.append({
                        "product_name": str(product_name).strip(),
                        "quantity": quantity, "unit": str(unit).strip(),
                        "unit_price": unit_price, "total_price": total_price,
                        "category_name": str(category_name).strip() if category_name else "其他",
                        "supplier_name": str(supplier_name).strip(),
                        "batch_number": str(batch_number).strip(),
                        "production_date": str(production_date).strip() if production_date else "",
                        "expiry_date": expiry_date,
                        "operator": str(operator).strip() if operator else "",
                        "stockin_date": stockin_date,
                        "stockin_created_at": stockin_created_at,
                    })
                except Exception:
                    pass

            progress.setValue(total_rows)
            if not parsed_data:
                return False, "Excel文件中没有有效数据"

            # 同日同食材汇总
            progress.setLabelText("正在汇总同文件数据...")
            grouped = defaultdict(lambda: {"quantity": 0, "total_price": 0, "items": []})
            for item in parsed_data:
                key = (item["product_name"], item["stockin_date"])
                grouped[key]["quantity"] += item["quantity"]
                grouped[key]["total_price"] += item["total_price"]
                grouped[key]["items"].append(item)

            summary_data = []
            for (name, date), group in grouped.items():
                first = group["items"][0]
                avg_price = group["total_price"] / group["quantity"] if group["quantity"] > 0 else first["unit_price"]
                summary_data.append({
                    **first, "product_name": name,
                    "quantity": group["quantity"],
                    "unit_price": avg_price,
                    "total_price": group["total_price"],
                    "stockin_date": date,
                })

            progress.setValue(total_rows + 5)
            progress.setLabelText("正在导入数据...")

            success, duplicate, error = 0, 0, 0
            errors = []
            for idx, item in enumerate(summary_data):
                if progress.wasCanceled():
                    return False, "用户取消导入"
                progress.setValue(total_rows + 10 + idx)
                try:
                    result = stock_service.import_stock_in(
                        product_name=item["product_name"],
                        quantity=item["quantity"],
                        unit_price=item["unit_price"],
                        unit=item["unit"],
                        category_name=item["category_name"],
                        supplier_name=item["supplier_name"],
                        batch_number=item["batch_number"],
                        production_date=item["production_date"],
                        expiry_date=item["expiry_date"],
                        operator=item["operator"],
                        created_at=item["stockin_created_at"],
                        remark="从销售订单导入",
                    )
                    if result is None:
                        # 服务层检测到完全相同的历史导入记录，跳过
                        duplicate += 1
                    else:
                        success += 1
                except Exception as e:
                    error += 1
                    errors.append(f"{item['product_name']}: {str(e)}")

            progress.close()
            msg = (f"导入完成!\nExcel解析行数: {len(parsed_data)} 行\n"
                   f"汇总后食材数: {len(summary_data)} 种\n"
                   f"新增: {success} 条\n"
                   f"重复跳过: {duplicate} 条\n"
                   f"失败: {error} 条")
            if duplicate:
                msg += ("\n\n检测到重复数据（相同食材/数量/单价/发货日期/单据编号），"
                        "已自动跳过，不会重复累加库存。\n"
                        "如需再次导入，请先删除对应的历史入库记录。")
            if errors:
                msg += "\n\n部分错误:\n" + "\n".join(errors[:5])
            return True, msg
        except Exception as e:
            import traceback
            return False, f"导入失败: {str(e)}\n{traceback.format_exc()}"

    @staticmethod
    def import_inspection_records(parent, stock_service, inspection_service=None, file_path: str = None):
        """从进货查验记录表导入数据（兼容 V1：自动建食材并生成入库记录，同时写入查验记录）"""
        if not file_path:
            file_path, _ = QFileDialog.getOpenFileName(
                parent, "导入进货查验记录", "", "Excel Files (*.xlsx)")
        if not file_path:
            return False, "未选择文件"

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            # 模板为两行合并表头，拼接成 "主_副" 形式
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
                headers.append([str(cell).strip() if cell else "" for cell in row])
            col_count = max(len(h) for h in headers)
            combined_headers = []
            for i in range(col_count):
                h1 = headers[0][i] if i < len(headers[0]) else ""
                h2 = headers[1][i] if i < len(headers[1]) else ""
                if h2 and h2 not in ["None", ""]:
                    combined_headers.append(f"{h1}_{h2}" if h1 else h2)
                else:
                    combined_headers.append(h1)
            header_map = {}
            for i, header in enumerate(combined_headers):
                header_map[header] = i

            success, error = 0, 0
            errors = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
                try:
                    product_name = row[header_map.get("产品名称", 1)]
                    if not product_name:
                        continue

                    quantity = row[header_map.get("数量", 3)]
                    unit = row[header_map.get("单位", 2)]
                    production_date = row[header_map.get("生产日期", 4)]
                    supplier_name = row[header_map.get("供货单位", 7)]
                    operator = row[header_map.get("记录人", 15)] if "记录人" in header_map else ""

                    if quantity is None or unit is None:
                        continue

                    quantity = float(quantity)
                    stock_in = stock_service.import_stock_in(
                        product_name=str(product_name).strip(),
                        quantity=quantity,
                        unit_price=0,
                        unit=str(unit).strip(),
                        category_name="",
                        supplier_name=str(supplier_name).strip() if supplier_name else "",
                        production_date=str(production_date).strip() if production_date else "",
                        operator=str(operator).strip() if operator else "",
                        remark="从进货查验记录导入",
                    )
                    if stock_in is None:
                        # 完全相同的查验导入记录已存在，跳过（入库与查验记录都不重复写入）
                        continue
                    # 同步写入进货查验记录（此前只建了入库，导致查验模块缺数据）
                    if inspection_service is not None:
                        try:
                            inspection_service.upsert_record(
                                stock_in_id=stock_in.id,
                                ingredient_id=stock_in.ingredient_id,
                                quantity=quantity, unit=str(unit).strip(),
                                production_date=str(production_date).strip() if production_date else "",
                                shelf_life="",
                                supplier_name=str(supplier_name).strip() if supplier_name else "",
                                batch_number="",
                                inspection_result="",
                                inspector=str(operator).strip() if operator else "",
                                inspection_date=str(production_date).strip() if production_date else "",
                                certificate_no="",
                                remark="从进货查验记录导入",
                            )
                        except Exception as e:
                            errors.append(f"第{row_idx}行 查验记录写入失败: {e}")
                    success += 1
                except Exception as e:
                    error += 1
                    errors.append(f"第{row_idx}行: {str(e)}")

            msg = f"导入完成!\n成功: {success} 条\n失败: {error} 条"
            if errors:
                msg += "\n\n部分错误:\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    msg += f"\n... 还有 {len(errors) - 5} 条"
            return True, msg
        except Exception as e:
            import traceback
            return False, f"导入失败: {str(e)}\n{traceback.format_exc()}"

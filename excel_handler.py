"""
Excel 导入导出处理模块 - macOS 风格
"""
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import QFileDialog, QMessageBox

from database import (CategoryDAO, SupplierDAO, IngredientDAO,
                      StockInDAO, StockOutDAO, CategoryMappingDAO)

DEFAULT_CATEGORY_MAP = {
    '鸡肉类': '肉类',
    '蛋类': '蛋类',
    '蔬菜瓜果类': '蔬菜类',
    '干货类': '粮油类',
    '豆制品': '豆制品',
    '调味品': '调味品',
    '水产类': '水产类',
    '水果类': '水果类',
    '猪肉类': '肉类',
    '牛肉类': '肉类',
    '粮油类': '粮油类',
}


def get_category_map():
    """获取类别映射（优先从数据库，无则使用默认）"""
    mappings = CategoryMappingDAO.get_all()
    if mappings:
        return {m.source_category: m.target_category_name for m in mappings}
    return DEFAULT_CATEGORY_MAP


# macOS 风格配色
MAC_BLUE = "0071e3"
MAC_GREEN = "34c759"
MAC_RED = "ff3b30"
MAC_ORANGE = "ff9500"
MAC_GRAY = "86868b"
MAC_LIGHT_GRAY = "f5f5f7"
MAC_BORDER = "d2d2d7"


class ExcelHandler:
    """Excel 处理类"""

    @staticmethod
    def _create_header_style():
        """创建 macOS 风格表头样式"""
        return {
            'fill': PatternFill(start_color=MAC_LIGHT_GRAY, end_color=MAC_LIGHT_GRAY, fill_type="solid"),
            'font': Font(color="1d1d1f", bold=True, size=12, name="-apple-system"),
            'alignment': Alignment(horizontal="left", vertical="center"),
            'border': Border(
                bottom=Side(style='thin', color=MAC_BORDER)
            )
        }

    @staticmethod
    def _create_cell_style():
        """创建 macOS 风格单元格样式"""
        return {
            'font': Font(color="1d1d1f", size=11, name="-apple-system"),
            'alignment': Alignment(horizontal="left", vertical="center"),
            'border': Border(
                bottom=Side(style='thin', color='f0f0f0')
            )
        }

    @staticmethod
    def export_ingredients(parent_widget, file_path=None):
        """导出食材余量信息到 Excel"""
        if not file_path:
            file_path, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "导出食材余量信息",
                f"食材余量_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )

        if not file_path:
            return False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "食材余量"

            header_style = ExcelHandler._create_header_style()
            cell_style = ExcelHandler._create_cell_style()

            # 写入表头
            headers = ["ID", "食材名称", "分类", "规格", "单位", "当前库存", "安全库存", "库存状态", "供应商"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_style['fill']
                cell.font = header_style['font']
                cell.alignment = header_style['alignment']
                cell.border = header_style['border']

            # 获取数据
            ingredients = IngredientDAO.get_all()

            # 写入数据
            for row, ing in enumerate(ingredients, 2):
                status = "正常" if ing.current_stock > ing.safety_stock else "库存不足"

                data = [
                    ing.id, ing.name, ing.category_name, ing.specification,
                    ing.unit, ing.current_stock, ing.safety_stock, status, ing.supplier_name
                ]

                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = cell_style['font']
                    cell.alignment = cell_style['alignment']
                    cell.border = cell_style['border']

                # 库存不足高亮
                if ing.current_stock <= ing.safety_stock:
                    for col in range(1, 10):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="fff5f5", end_color="fff5f5", fill_type="solid"
                        )
                        ws.cell(row=row, column=col).font = Font(
                            color=MAC_RED, size=11, name="-apple-system"
                        )

            # 调整列宽
            column_widths = [8, 22, 14, 16, 10, 14, 14, 14, 18]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # 添加汇总信息
            summary_row = len(ingredients) + 3
            ws.cell(row=summary_row, column=1, value="汇总信息")
            ws.cell(row=summary_row, column=1).font = Font(
                bold=True, size=13, color=MAC_BLUE, name="-apple-system"
            )

            ws.cell(row=summary_row + 1, column=1, value="食材种类总数:")
            ws.cell(row=summary_row + 1, column=2, value=len(ingredients))

            low_stock_count = sum(1 for ing in ingredients if ing.current_stock <= ing.safety_stock)
            ws.cell(row=summary_row + 2, column=1, value="库存预警数量:")
            ws.cell(row=summary_row + 2, column=2, value=low_stock_count)
            ws.cell(row=summary_row + 2, column=2).font = Font(
                color=MAC_RED, bold=True, name="-apple-system"
            )

            wb.save(file_path)
            return True

        except Exception as e:
            QMessageBox.critical(parent_widget, "导出失败", f"导出时发生错误:\n{str(e)}")
            return False

    @staticmethod
    def import_ingredients(parent_widget, file_path=None):
        """从 Excel 导入食材信息"""
        if not file_path:
            file_path, _ = QFileDialog.getOpenFileName(
                parent_widget,
                "导入食材信息",
                "",
                "Excel Files (*.xlsx)"
            )

        if not file_path:
            return False, "未选择文件"

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            # 验证表头
            expected_headers = ["食材名称", "分类", "规格", "单位", "安全库存", "供应商"]
            actual_headers = [cell.value for cell in ws[1]]

            # 查找必需列的索引
            header_map = {}
            for i, header in enumerate(actual_headers):
                if header in expected_headers:
                    header_map[header] = i

            if "食材名称" not in header_map or "单位" not in header_map:
                return False, "Excel 格式不正确，必须包含'食材名称'和'单位'列"

            # 获取分类和供应商映射
            categories = {cat.name: cat.id for cat in CategoryDAO.get_all()}
            suppliers = {s.name: s.id for s in SupplierDAO.get_all()}

            # 获取已有食材名称集合，用于查重
            existing_ingredients = {ing.name.lower() for ing in IngredientDAO.get_all()}
            imported_names = set()  # 本次导入中的名称，防止文件内重复

            success_count = 0
            error_count = 0
            skip_count = 0
            new_categories = []
            new_suppliers = []
            errors = []

            # 从第2行开始读取数据
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    name = row[header_map.get("食材名称", 0)]
                    if not name:
                        continue

                    name_str = str(name).strip()

                    # 查重：跳过已存在的食材
                    if name_str.lower() in existing_ingredients:
                        skip_count += 1
                        errors.append(f"第{row_idx}行: 食材'{name_str}'已存在，跳过")
                        continue
                    # 文件内重复
                    if name_str.lower() in imported_names:
                        skip_count += 1
                        errors.append(f"第{row_idx}行: 食材'{name_str}'在文件内重复，跳过")
                        continue
                    imported_names.add(name_str.lower())

                    category_name = row[header_map.get("分类", 1)] if "分类" in header_map else None
                    specification = row[header_map.get("规格", 2)] if "规格" in header_map else ""
                    unit = row[header_map.get("单位", 3)] if "单位" in header_map else ""

                    safety_stock = 0
                    if "安全库存" in header_map:
                        try:
                            safety_stock = float(row[header_map["安全库存"]])
                        except (ValueError, TypeError):
                            safety_stock = 0

                    supplier_name = row[header_map.get("供应商", 5)] if "供应商" in header_map else None

                    # 处理分类
                    category_id = None
                    if category_name:
                        if category_name not in categories:
                            CategoryDAO.add(category_name, f"从Excel导入: {category_name}")
                            categories = {cat.name: cat.id for cat in CategoryDAO.get_all()}
                            new_categories.append(category_name)
                        category_id = categories.get(category_name)

                    # 处理供应商
                    supplier_id = None
                    if supplier_name:
                        if supplier_name not in suppliers:
                            SupplierDAO.add(supplier_name)
                            suppliers = {s.name: s.id for s in SupplierDAO.get_all()}
                            new_suppliers.append(supplier_name)
                        supplier_id = suppliers.get(supplier_name)

                    # 添加食材
                    IngredientDAO.add(
                        name=name_str,
                        category_id=category_id,
                        unit=str(unit) if unit else "个",
                        specification=str(specification) if specification else "",
                        safety_stock=safety_stock,
                        supplier_id=supplier_id
                    )
                    existing_ingredients.add(name_str.lower())
                    success_count += 1

                except Exception as e:
                    error_count += 1
                    errors.append(f"第{row_idx}行: {str(e)}")

            result_msg = f"导入完成!\n成功: {success_count} 条\n跳过(重复): {skip_count} 条\n失败: {error_count} 条"
            if new_categories:
                result_msg += f"\n\n自动创建分类({len(new_categories)}个): {', '.join(new_categories[:5])}"
                if len(new_categories) > 5:
                    result_msg += f" 等{len(new_categories)}个"
            if new_suppliers:
                result_msg += f"\n自动创建供应商({len(new_suppliers)}个): {', '.join(new_suppliers[:5])}"
                if len(new_suppliers) > 5:
                    result_msg += f" 等{len(new_suppliers)}个"
            if errors:
                result_msg += "\n\n详情:\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    result_msg += f"\n... 还有 {len(errors) - 5} 条"

            return True, result_msg

        except Exception as e:
            return False, f"导入失败: {str(e)}"

    @staticmethod
    def create_template(parent_widget):
        """创建 macOS 风格导入模板"""
        file_path, _ = QFileDialog.getSaveFileName(
            parent_widget,
            "保存导入模板",
            "食材导入模板.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return False

        try:
            wb = Workbook()

            # === 主模板页 ===
            ws = wb.active
            ws.title = "食材导入模板"

            # 标题行
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = "学校食堂食材导入模板"
            title_cell.font = Font(size=18, bold=True, color=MAC_BLUE, name="-apple-system")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 36

            # 副标题
            ws.merge_cells('A2:F2')
            subtitle_cell = ws['A2']
            subtitle_cell.value = "请按照下方格式填写食材信息，带 * 号为必填项"
            subtitle_cell.font = Font(size=11, color=MAC_GRAY, name="-apple-system")
            subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 24

            # 表头 (第4行)
            headers = [
                ("食材名称*", "食材的完整名称，如：大白菜、五花肉"),
                ("分类", "所属分类，如：蔬菜类、肉类、粮油类"),
                ("规格", "规格描述，如：新鲜、精品、散装"),
                ("单位*", "计量单位，如：斤、个、袋、桶"),
                ("安全库存", "库存预警阈值，低于此值会提醒补货"),
                ("供应商", "供应商名称，不存在会自动创建")
            ]

            header_fill = PatternFill(start_color=MAC_LIGHT_GRAY, end_color=MAC_LIGHT_GRAY, fill_type="solid")
            header_font = Font(color="1d1d1f", bold=True, size=12, name="-apple-system")
            header_border = Border(bottom=Side(style='medium', color=MAC_BORDER))

            for col, (header, _) in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = header_border
                cell.alignment = Alignment(horizontal="left", vertical="center")

            ws.row_dimensions[4].height = 28

            # 示例数据
            examples = [
                ["大白菜", "蔬菜类", "新鲜", "斤", 50, "绿源蔬菜批发"],
                ["五花肉", "肉类", "精品", "斤", 30, "鸿运肉业"],
                ["鸡蛋", "蛋类", "散养土鸡蛋", "个", 200, "阳光养殖场"],
                ["东北大米", "粮油类", "五常稻花香", "袋", 20, "金穗粮油"],
                ["食用油", "粮油类", "非转基因压榨", "桶", 10, "金穗粮油"],
                ["西红柿", "蔬菜类", "有机", "斤", 40, "绿源蔬菜批发"],
                ["鸡胸肉", "肉类", "冷冻", "斤", 25, "鸿运肉业"],
                ["豆腐", "豆制品", "嫩豆腐", "块", 30, "豆香坊"],
            ]

            cell_font = Font(size=11, name="-apple-system")
            for row_idx, example in enumerate(examples, 5):
                for col_idx, value in enumerate(example, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = cell_font
                    cell.border = Border(bottom=Side(style='thin', color='f0f0f0'))
                ws.row_dimensions[row_idx].height = 24

            # 调整列宽
            column_widths = [18, 14, 18, 10, 14, 18]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # === 填写说明页 ===
            ws2 = wb.create_sheet("填写说明")

            # 标题
            ws2.merge_cells('A1:C1')
            ws2['A1'] = "填写说明"
            ws2['A1'].font = Font(size=16, bold=True, color=MAC_BLUE, name="-apple-system")
            ws2['A1'].alignment = Alignment(horizontal="left", vertical="center")
            ws2.row_dimensions[1].height = 32

            # 表头说明表格
            ws2['A3'] = "字段名"
            ws2['B3'] = "说明"
            ws2['C3'] = "是否必填"
            for col in range(1, 4):
                cell = ws2.cell(row=3, column=col)
                cell.font = Font(bold=True, size=12, name="-apple-system")
                cell.fill = PatternFill(start_color=MAC_LIGHT_GRAY, end_color=MAC_LIGHT_GRAY, fill_type="solid")
                cell.border = Border(bottom=Side(style='medium', color=MAC_BORDER))

            field_details = [
                ("食材名称", "食材的完整名称，建议简洁明了", "是"),
                ("分类", "食材所属分类。常见分类：蔬菜类、肉类、水产类、豆制品、粮油类、调味品、蛋类、水果类。不存在的分类会自动创建", "否"),
                ("规格", "食材的规格描述，如产地、等级、包装方式等", "否"),
                ("单位", "计量单位，如：斤、公斤、个、袋、桶、箱等", "是"),
                ("安全库存", "库存预警阈值。当库存低于此值时系统会提醒补货。不填默认为0", "否"),
                ("供应商", "供应商名称。不存在的供应商会自动创建", "否"),
            ]

            for row_idx, (field, desc, required) in enumerate(field_details, 4):
                ws2.cell(row=row_idx, column=1, value=field).font = Font(
                    bold=True, size=11, name="-apple-system"
                )
                ws2.cell(row=row_idx, column=2, value=desc).font = Font(
                    size=11, name="-apple-system"
                )
                req_cell = ws2.cell(row=row_idx, column=3, value=required)
                req_cell.font = Font(
                    bold=True, size=11,
                    color=MAC_RED if required == "是" else MAC_GREEN,
                    name="-apple-system"
                )
                req_cell.alignment = Alignment(horizontal="center")
                ws2.row_dimensions[row_idx].height = 28

            # 注意事项
            notice_row = len(field_details) + 6
            ws2.merge_cells(f'A{notice_row}:C{notice_row}')
            ws2.cell(row=notice_row, column=1, value="注意事项").font = Font(
                size=14, bold=True, color=MAC_ORANGE, name="-apple-system"
            )

            notices = [
                "1. 第一行（第4行）为表头，请勿删除或修改",
                "2. 从第5行开始填写实际数据",
                "3. 分类和供应商如果不存在，导入时会自动创建",
                "4. 安全库存请填写数字，不填默认为0",
                "5. 建议先下载此模板，按格式填写后再导入",
                "6. 导入前请确保数据准确，避免重复导入",
            ]

            for i, notice in enumerate(notices, notice_row + 1):
                ws2.merge_cells(f'A{i}:C{i}')
                cell = ws2.cell(row=i, column=1, value=notice)
                cell.font = Font(size=11, name="-apple-system")
                ws2.row_dimensions[i].height = 24

            ws2.column_dimensions['A'].width = 16
            ws2.column_dimensions['B'].width = 55
            ws2.column_dimensions['C'].width = 12

            # === 分类参考页 ===
            ws3 = wb.create_sheet("分类参考")
            ws3['A1'] = "系统默认分类"
            ws3['A1'].font = Font(size=14, bold=True, color=MAC_BLUE, name="-apple-system")

            default_categories = [
                "蔬菜类", "肉类", "水产类", "豆制品",
                "粮油类", "调味品", "蛋类", "水果类"
            ]

            for i, cat in enumerate(default_categories, 3):
                ws3.cell(row=i, column=1, value=cat).font = Font(
                    size=11, name="-apple-system"
                )

            ws3.column_dimensions['A'].width = 16

            wb.save(file_path)
            return True

        except Exception as e:
            QMessageBox.critical(parent_widget, "创建模板失败", str(e))
            return False

    @staticmethod
    def export_stock_records(parent_widget, record_type="in"):
        """导出出入库记录"""
        file_path, _ = QFileDialog.getSaveFileName(
            parent_widget,
            f"导出{'入库' if record_type == 'in' else '出库'}记录",
            f"{'入库' if record_type == 'in' else '出库'}记录_{datetime.now().strftime('%Y%m%d')}.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "记录"

            header_style = ExcelHandler._create_header_style()
            cell_style = ExcelHandler._create_cell_style()

            if record_type == "in":
                headers = ["ID", "食材", "数量", "单价", "总价", "供应商", "批次号", "生产日期", "保质期", "操作人", "备注", "时间"]
                records = StockInDAO.get_all(1000)

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_style['fill']
                    cell.font = header_style['font']
                    cell.border = header_style['border']

                for row, r in enumerate(records, 2):
                    data = [r.id, r.ingredient_name, r.quantity, r.unit_price,
                           r.total_price, r.supplier_id, r.batch_number,
                           r.production_date, r.expiry_date, r.operator, r.remark, r.created_at]
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = cell_style['font']
                        cell.border = cell_style['border']
            else:
                headers = ["ID", "食材", "数量", "单价", "总价", "用途", "领用部门", "操作人", "备注", "时间"]
                records = StockOutDAO.get_all(1000)

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_style['fill']
                    cell.font = header_style['font']
                    cell.border = header_style['border']

                for row, r in enumerate(records, 2):
                    data = [r.id, r.ingredient_name, r.quantity, r.unit_price,
                           r.total_price, r.purpose, r.department, r.operator, r.remark, r.created_at]
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = cell_style['font']
                        cell.border = cell_style['border']

            wb.save(file_path)
            return True

        except Exception as e:
            QMessageBox.critical(parent_widget, "导出失败", str(e))
            return False

    @staticmethod
    def import_sales_orders(parent_widget, file_path=None):
        """从销售订单Excel导入数据，支持合并单元格处理和重复检测"""
        if not file_path:
            file_path, _ = QFileDialog.getOpenFileName(
                parent_widget,
                "导入销售订单",
                "",
                "Excel Files (*.xlsx)"
            )

        if not file_path:
            return False, "未选择文件"

        try:
            from PyQt6.QtWidgets import QProgressDialog, QDialog, QVBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QPushButton, QHBoxLayout, QMessageBox, QHeaderView
            from PyQt6.QtCore import Qt
            from datetime import datetime as dt, timedelta
            
            progress = QProgressDialog("正在解析Excel文件...", "取消", 0, 100, parent_widget)
            progress.setWindowTitle("导入销售订单")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(10)
            
            wb = load_workbook(file_path)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            header_map = {}
            for i, header in enumerate(headers):
                header_map[str(header).strip() if header else ""] = i

            required_headers = ['商品名称', '数量', '单位', '含税单价', '价税合计']
            missing = [h for h in required_headers if h not in header_map]
            if missing:
                return False, f"缺少必需列: {', '.join(missing)}"

            progress.setValue(15)
            progress.setLabelText("正在处理合并单元格...")

            # 获取所有行数据（不使用values_only，以便处理合并单元格）
            all_rows = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
                all_rows.append(row)

            # 构建合并单元格映射：对于每个合并区域，将值填充到所有单元格
            merged_fill = {}
            for merged_range in ws.merged_cells.ranges:
                min_col = merged_range.min_col - 1
                min_row = merged_range.min_row - 1
                max_col = merged_range.max_col - 1
                max_row = merged_range.max_row - 1
                top_left_value = all_rows[min_row][min_col].value
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        merged_fill[(r, c)] = top_left_value

            # 获取所有行的值（填充合并单元格）
            row_values = []
            for row_idx, row in enumerate(all_rows):
                vals = []
                for col_idx, cell in enumerate(row):
                    if (row_idx, col_idx) in merged_fill:
                        vals.append(merged_fill[(row_idx, col_idx)])
                    else:
                        vals.append(cell.value)
                row_values.append(vals)

            progress.setValue(20)
            progress.setLabelText("正在解析数据...")
            
            categories = {cat.name: cat.id for cat in CategoryDAO.get_all()}
            suppliers = {s.name: s.id for s in SupplierDAO.get_all()}
            category_map = get_category_map()

            # 解析所有数据行（从第2行开始，第1行是表头）
            parsed_data = []
            total_rows = len(row_values) - 1
            
            progress.setLabelText(f"正在解析数据（共{total_rows}行）...")
            progress.setMaximum(total_rows + 20)

            def parse_date(val):
                """解析日期，返回date字符串和datetime字符串"""
                if not val:
                    return "", ""
                if hasattr(val, 'strftime'):
                    return val.strftime("%Y-%m-%d"), val.strftime("%Y-%m-%d 08:00:00")
                val_str = str(val).strip()
                if not val_str:
                    return "", ""
                # 尝试多种日期格式
                for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y",
                            "%Y.%m.%d", "%Y年%m月%d日", "%Y-%m-%d %H:%M:%S",
                            "%Y/%m/%d %H:%M:%S", "%m-%d-%Y", "%d-%m-%Y"]:
                    try:
                        d = dt.strptime(val_str, fmt)
                        return d.strftime("%Y-%m-%d"), d.strftime("%Y-%m-%d 08:00:00")
                    except:
                        continue
                # 尝试提取日期部分（如 "2026/6/20 08:00:00" 格式）
                import re
                m = re.match(r'(\d{4})[/.年](\d{1,2})[/.月](\d{1,2})', val_str)
                if m:
                    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
                    return f"{y:04d}-{mo:02d}-{d:02d}", f"{y:04d}-{mo:02d}-{d:02d} 08:00:00"
                return "", ""
            
            for i in range(1, len(row_values)):
                if progress.wasCanceled():
                    return False, "用户取消导入"
                
                progress.setValue(i)
                row = row_values[i]
                row_idx = i + 1
                
                try:
                    product_name = row[header_map.get('商品名称', 0)]
                    if not product_name:
                        continue

                    quantity = row[header_map.get('数量', 0)]
                    unit = row[header_map.get('单位', 0)]
                    unit_price = row[header_map.get('含税单价', 0)]
                    total_price = row[header_map.get('价税合计', 0)]

                    if quantity is None or unit is None or unit_price is None:
                        continue

                    quantity = float(quantity)
                    unit_price = float(unit_price)

                    if total_price is None:
                        total_price = quantity * unit_price
                    else:
                        total_price = float(total_price)

                    category_name = row[header_map.get('商品类别', 0)]
                    if category_name:
                        category_name = category_map.get(str(category_name).strip(), str(category_name).strip())
                    else:
                        category_name = "其他"

                    supplier_name = row[header_map.get('生产单位（空白部分自己查验收单）', 0)]
                    if not supplier_name:
                        supplier_name = "未知供应商"
                    else:
                        supplier_name = str(supplier_name).strip()

                    # 发货日期作为入库日期
                    delivery_date = row[header_map.get('发货日期', 0)]
                    stockin_date, stockin_created_at = parse_date(delivery_date)
                    
                    if not stockin_date:
                        stockin_date = dt.now().strftime("%Y-%m-%d")
                        stockin_created_at = dt.now().strftime("%Y-%m-%d 08:00:00")

                    batch_number = row[header_map.get('单据编号', 0)]
                    production_date = row[header_map.get('生鲜日期', 0)]
                    expiry_days = row[header_map.get('保质期(天)', 0)]
                    operator = row[header_map.get('业务员', 0)]
                    
                    # 计算到期日期
                    expiry_date = ""
                    if production_date and expiry_days:
                        try:
                            prod_date_str, _ = parse_date(production_date)
                            if prod_date_str:
                                prod_d = dt.strptime(prod_date_str, "%Y-%m-%d")
                                expiry_days_int = int(float(expiry_days))
                                expiry_date = (prod_d + timedelta(days=expiry_days_int)).strftime("%Y-%m-%d")
                        except:
                            pass
                    
                    parsed_data.append({
                        'row_idx': row_idx,
                        'product_name': str(product_name).strip(),
                        'quantity': quantity,
                        'unit': str(unit).strip(),
                        'unit_price': unit_price,
                        'total_price': total_price,
                        'category_name': category_name,
                        'supplier_name': supplier_name,
                        'batch_number': str(batch_number).strip() if batch_number else "",
                        'production_date': str(production_date).strip() if production_date else "",
                        'expiry_date': expiry_date,
                        'operator': str(operator).strip() if operator else "",
                        'stockin_date': stockin_date,
                        'stockin_created_at': stockin_created_at,
                        'remark': "从销售订单导入"
                    })

                except Exception as e:
                    pass
            
            progress.setValue(total_rows)
            
            if not parsed_data:
                return False, "Excel文件中没有有效数据"

            # 同文件内相同食材+同一天的数据汇总
            progress.setLabelText("正在汇总同文件数据...")
            from collections import defaultdict
            
            grouped = defaultdict(lambda: {'quantity': 0, 'total_price': 0, 'items': []})
            for item in parsed_data:
                key = (item['product_name'], item['stockin_date'])
                grouped[key]['quantity'] += item['quantity']
                grouped[key]['total_price'] += item['total_price']
                grouped[key]['items'].append(item)
            
            # 构建汇总后的数据
            summary_data = []
            for (name, date), group in grouped.items():
                # 使用第一条的大部分字段
                first = group['items'][0]
                avg_price = group['total_price'] / group['quantity'] if group['quantity'] > 0 else first['unit_price']
                summary_data.append({
                    'row_idx': first['row_idx'],
                    'product_name': name,
                    'quantity': group['quantity'],
                    'unit': first['unit'],
                    'unit_price': avg_price,
                    'total_price': group['total_price'],
                    'category_name': first['category_name'],
                    'supplier_name': first['supplier_name'],
                    'batch_number': first['batch_number'],
                    'production_date': first['production_date'],
                    'expiry_date': first['expiry_date'],
                    'operator': first['operator'],
                    'stockin_date': date,
                    'stockin_created_at': first['stockin_created_at'],
                    'remark': first['remark'],
                    'original_count': len(group['items'])
                })
            
            progress.setValue(total_rows + 3)

            # 检测重复：当天已有相同食材的入库记录
            from database import get_connection
            duplicates = []
            unique_items = []
            
            progress.setLabelText("正在检测重复记录...")
            
            for item in summary_data:
                with get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT id, quantity, unit_price, total_price, supplier_id, batch_number
                        FROM stock_in
                        WHERE ingredient_id IN (SELECT id FROM ingredients WHERE name = ?)
                        AND strftime('%Y-%m-%d', created_at) = ?
                        AND remark LIKE '%从销售订单导入%'
                    ''', (item['product_name'], item['stockin_date']))
                    existing = cursor.fetchone()
                    
                    if existing:
                        duplicates.append({
                            'name': item['product_name'],
                            'date': item['stockin_date'],
                            'new_quantity': round(item['quantity'], 2),
                            'new_price': round(item['unit_price'], 2),
                            'existing_quantity': existing['quantity'],
                            'existing_price': existing['unit_price'],
                            'existing_id': existing['id'],
                            'item': item
                        })
                    else:
                        unique_items.append(item)
            
            progress.setValue(total_rows + 5)
            
            # 如果有重复，显示确认对话框
            action = None
            if duplicates:
                dup_dialog = QDialog(parent_widget)
                dup_dialog.setWindowTitle("发现重复记录")
                dup_dialog.setMinimumSize(700, 450)
                dup_dialog.setStyleSheet("""
                    QDialog { background-color: white; }
                    QLabel { font-size: 13px; color: #1d1d1f; }
                    QPushButton { background-color: #0071e3; color: white; border: none; padding: 8px 20px; border-radius: 6px; }
                    QPushButton:hover { background-color: #0077ed; }
                    QPushButton#skip { background-color: #86868b; }
                    QPushButton#cancel { background-color: #ff3b30; }
                    QTableWidget { border: 1px solid #d2d2d7; border-radius: 8px; }
                """)
                
                layout = QVBoxLayout(dup_dialog)
                
                info_label = QLabel(f"检测到以下 {len(duplicates)} 条记录在当天已有入库数据（来源相同）：\n请选择处理方式")
                layout.addWidget(info_label)
                
                dup_table = QTableWidget()
                dup_table.setColumnCount(6)
                dup_table.setHorizontalHeaderLabels(["食材名称", "日期", "新数量", "现有数量", "新单价", "现有单价"])
                dup_table.setRowCount(len(duplicates))
                
                for i, dup in enumerate(duplicates):
                    dup_table.setItem(i, 0, QTableWidgetItem(dup['name']))
                    dup_table.setItem(i, 1, QTableWidgetItem(dup['date']))
                    dup_table.setItem(i, 2, QTableWidgetItem(str(dup['new_quantity'])))
                    dup_table.setItem(i, 3, QTableWidgetItem(str(dup['existing_quantity'])))
                    dup_table.setItem(i, 4, QTableWidgetItem(f"¥{dup['new_price']}"))
                    dup_table.setItem(i, 5, QTableWidgetItem(f"¥{dup['existing_price']}"))
                
                dup_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
                dup_table.setAlternatingRowColors(True)
                layout.addWidget(dup_table)
                
                btn_layout = QHBoxLayout()
                
                btn_add = QPushButton("全部新增")
                btn_add.setToolTip("将重复记录也作为新记录导入")
                btn_layout.addWidget(btn_add)
                
                btn_update = QPushButton("全部覆盖")
                btn_update.setToolTip("用新数据覆盖当天已有的记录")
                btn_update.setStyleSheet("background-color: #ff9500;")
                btn_layout.addWidget(btn_update)
                
                btn_skip = QPushButton("跳过重复")
                btn_skip.setObjectName("skip")
                btn_skip.setToolTip("只导入不重复的记录，跳过重复项")
                btn_layout.addWidget(btn_skip)
                
                btn_cancel = QPushButton("取消导入")
                btn_cancel.setObjectName("cancel")
                btn_layout.addWidget(btn_cancel)
                
                layout.addLayout(btn_layout)
                
                action_holder = [None]
                
                def on_add():
                    action_holder[0] = 'add'
                    dup_dialog.accept()
                
                def on_update():
                    action_holder[0] = 'update'
                    dup_dialog.accept()
                
                def on_skip():
                    action_holder[0] = 'skip'
                    dup_dialog.accept()
                
                def on_cancel():
                    action_holder[0] = 'cancel'
                    dup_dialog.reject()
                
                btn_add.clicked.connect(on_add)
                btn_update.clicked.connect(on_update)
                btn_skip.clicked.connect(on_skip)
                btn_cancel.clicked.connect(on_cancel)
                
                dup_dialog.exec()
                
                action = action_holder[0]
                
                if action == 'cancel' or not action:
                    return False, "用户取消导入"
                
                if action == 'add':
                    import_items = summary_data
                elif action == 'skip':
                    import_items = unique_items
                elif action == 'update':
                    import_items = summary_data
            else:
                import_items = summary_data
            
            progress.setLabelText("正在导入数据...")
            progress.setMaximum(len(import_items) + total_rows + 10)
            
            success_count = 0
            update_count = 0
            error_count = 0
            errors = []
            
            for idx, item in enumerate(import_items):
                if progress.wasCanceled():
                    return False, "用户取消导入"
                
                progress.setValue(total_rows + 10 + idx)
                
                try:
                    category_name = item['category_name']
                    if category_name not in categories:
                        CategoryDAO.add(category_name, f"从销售订单导入: {category_name}")
                        categories = {cat.name: cat.id for cat in CategoryDAO.get_all()}
                    category_id = categories.get(category_name)

                    supplier_name = item['supplier_name']
                    if supplier_name and supplier_name not in suppliers:
                        SupplierDAO.add(supplier_name)
                        suppliers = {s.name: s.id for s in SupplierDAO.get_all()}
                    supplier_id = suppliers.get(supplier_name) if supplier_name else None

                    ingredients = IngredientDAO.get_all()
                    ing = next((i for i in ingredients if i.name == item['product_name']), None)
                    if not ing:
                        IngredientDAO.add(
                            name=item['product_name'],
                            category_id=category_id,
                            unit=item['unit'],
                            supplier_id=supplier_id
                        )
                        ingredients = IngredientDAO.get_all()
                        ing = next((i for i in ingredients if i.name == item['product_name']), None)

                    if ing:
                        # 检查是否是更新模式且存在重复
                        is_duplicate = any(d['item'] is item for d in duplicates)
                        
                        if action == 'update' and is_duplicate:
                            dup_entry = next(d for d in duplicates if d['item'] is item)
                            with get_connection() as conn:
                                cursor = conn.cursor()
                                cursor.execute('''
                                    UPDATE stock_in SET
                                        quantity = ?, unit_price = ?, total_price = ?,
                                        supplier_id = ?, batch_number = ?, production_date = ?,
                                        expiry_date = ?, operator = ?, remark = ?
                                    WHERE id = ?
                                ''', (item['quantity'], item['unit_price'], item['total_price'],
                                      supplier_id, item['batch_number'], item['production_date'],
                                      item['expiry_date'], item['operator'], item['remark'],
                                      dup_entry['existing_id']))
                                conn.commit()
                                update_count += 1
                        else:
                            StockInDAO.add(
                                ingredient_id=ing.id,
                                quantity=item['quantity'],
                                unit_price=item['unit_price'],
                                supplier_id=supplier_id,
                                batch_number=item['batch_number'],
                                production_date=item['production_date'],
                                expiry_date=item['expiry_date'],
                                operator=item['operator'],
                                remark=item['remark'],
                                created_at=item['stockin_created_at']
                            )
                            success_count += 1

                except Exception as e:
                    error_count += 1
                    errors.append(f"第{item['row_idx']}行: {str(e)}")

            progress.setValue(total_rows + 10 + len(import_items))
            progress.close()
            
            # 计算原始行数
            original_total = sum(item.get('original_count', 1) for item in import_items)
            
            result_msg = (f"导入完成!\n"
                         f"Excel解析行数: {len(parsed_data)} 行\n"
                         f"汇总后食材数: {len(summary_data)} 种\n"
                         f"新增: {success_count} 条\n"
                         f"更新: {update_count} 条\n"
                         f"失败: {error_count} 条")
            if errors:
                result_msg += "\n\n部分错误:\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    result_msg += f"\n... 还有 {len(errors) - 5} 条错误"

            return True, result_msg

        except Exception as e:
            import traceback
            return False, f"导入失败: {str(e)}\n{traceback.format_exc()}"

    @staticmethod
    def import_inspection_records(parent_widget, file_path=None):
        """从进货查验记录表导入数据"""
        if not file_path:
            file_path, _ = QFileDialog.getOpenFileName(
                parent_widget,
                "导入进货查验记录",
                "",
                "Excel Files (*.xlsx)"
            )

        if not file_path:
            return False, "未选择文件"

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            headers = []
            for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
                headers.append([str(cell).strip() if cell else "" for cell in row])

            col_count = max(len(h) for h in headers)
            combined_headers = []
            for i in range(col_count):
                h1 = headers[0][i] if i < len(headers[0]) else ""
                h2 = headers[1][i] if i < len(headers[1]) else ""
                if h2 and h2 not in ['None', '']:
                    combined_headers.append(f"{h1}_{h2}" if h1 else h2)
                else:
                    combined_headers.append(h1)

            header_map = {}
            for i, header in enumerate(combined_headers):
                header_map[header] = i

            categories = {cat.name: cat.id for cat in CategoryDAO.get_all()}
            suppliers = {s.name: s.id for s in SupplierDAO.get_all()}

            success_count = 0
            error_count = 0
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
                try:
                    product_name = row[header_map.get('产品名称', 1)]
                    if not product_name:
                        continue

                    quantity = row[header_map.get('数量', 3)]
                    unit = row[header_map.get('单位', 2)]
                    production_date = row[header_map.get('生产日期', 4)]
                    shelf_life = row[header_map.get('保质期', 5)]
                    supplier_name = row[header_map.get('供货单位', 7)]

                    if quantity is None or unit is None:
                        continue

                    quantity = float(quantity)
                    unit_price = 0

                    if supplier_name and supplier_name not in suppliers:
                        SupplierDAO.add(str(supplier_name))
                        suppliers = {s.name: s.id for s in SupplierDAO.get_all()}
                    supplier_id = suppliers.get(str(supplier_name)) if supplier_name else None

                    ingredients = IngredientDAO.get_all()
                    ing = next((i for i in ingredients if i.name == str(product_name)), None)
                    if not ing:
                        IngredientDAO.add(
                            name=str(product_name),
                            category_id=categories.get("其他", None),
                            unit=str(unit)
                        )
                        ingredients = IngredientDAO.get_all()
                        ing = next((i for i in ingredients if i.name == str(product_name)), None)

                    if ing:
                        StockInDAO.add(
                            ingredient_id=ing.id,
                            quantity=quantity,
                            unit_price=unit_price,
                            supplier_id=supplier_id,
                            production_date=str(production_date) if production_date else "",
                            expiry_date="",
                            operator=row[header_map.get('记录人', 15)] if '记录人' in header_map else "",
                            remark="从进货查验记录导入"
                        )
                        success_count += 1

                except Exception as e:
                    error_count += 1
                    errors.append(f"第{row_idx}行: {str(e)}")

            result_msg = f"导入完成!\n成功: {success_count} 条\n失败: {error_count} 条"
            if errors:
                result_msg += "\n\n部分错误:\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    result_msg += f"\n... 还有 {len(errors) - 5} 条错误"

            return True, result_msg

        except Exception as e:
            return False, f"导入失败: {str(e)}"

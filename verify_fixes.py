"""
End-to-end verification of the 6 V2 defect fixes.
Run with the project venv python. ASCII-only source to avoid GBK pitfalls.
"""
import sys, os, tempfile
from datetime import datetime, date

ROOT = r"C:\Users\lcy\work\school-canteen-management"
sys.path.insert(0, ROOT)

# Bind the peewee proxy to an in-memory DB BEFORE importing model-dependent code.
from peewee import SqliteDatabase
import school_canteen.data.database as dbmod
mem = SqliteDatabase(":memory:", pragmas={"foreign_keys": 1})
dbmod.db.initialize(mem)

from school_canteen.data.models import (
    User, Category, Supplier, Ingredient, StockIn, StockOut,
    InventoryCheck, OperationLog, InspectionRecord,
)
dbmod.db.connect(reuse_if_open=True)
dbmod.db.create_tables(
    [User, Category, Supplier, Ingredient, StockIn, StockOut,
     InventoryCheck, OperationLog, InspectionRecord], safe=True)

from school_canteen.data.repositories.stock_repository import (
    StockInRepository, StockOutRepository)
from school_canteen.data.repositories.inspection_repository import (
    InspectionRecordRepository)
from school_canteen.data.repositories.ingredient_repository import (
    IngredientRepository)
from school_canteen.data.repositories.catalog_repository import (
    CategoryRepository, SupplierRepository)
from school_canteen.data.repositories.stock_repository import (
    InventoryCheckRepository)
from school_canteen.data.repositories.report_repository import ReportRepository
from school_canteen.services.stock_service import StockService
from school_canteen.utils.report_generator import ReportGenerator
from school_canteen.core.session import Session

# ---- Seed data ----
cat = Category.create(name="Veg")
sup = Supplier.create(name="SupA")
ing = Ingredient.create(name="Cabbage", category=cat, unit="jin",
                        supplier=sup, current_stock=0)
TARGET = "2026-06-15"
si1 = StockIn.create(ingredient=ing, quantity=50, unit_price=2.0,
                     total_price=100.0, supplier=sup,
                     created_at=datetime(2026, 6, 15, 9, 0, 0))
StockIn.create(ingredient=ing, quantity=30, unit_price=3.0,
               total_price=90.0, supplier=sup,
               created_at=datetime(2026, 6, 15, 10, 0, 0))
StockOut.create(ingredient=ing, quantity=20, unit_price=2.5,
                total_price=50.0, created_at=datetime(2026, 6, 15, 14, 0, 0))
InspectionRecord.create(stock_in=si1, ingredient=ing, quantity=50,
                        unit="jin", inspection_result="OK",
                        inspector="Zhang", inspection_date=date(2026, 6, 15))
admin = User.create(username="admin", password_hash="x", salt="", role="admin")
OperationLog.create(user=admin, action="stock_in", target_type="stock_in",
                    target_id=si1.id)

in_repo = StockInRepository()
out_repo = StockOutRepository()
insp_repo = InspectionRecordRepository()
ing_repo = IngredientRepository()
cat_repo = CategoryRepository()
sup_repo = SupplierRepository()
inv_repo = InventoryCheckRepository()
svc = StockService(ing_repo, in_repo, out_repo, inv_repo, cat_repo,
                   sup_repo, log_repo=None)

# Direct .create() calls above bypass the service, so sync the ingredient
# stock to reflect the seeded movements: 0 +50 +30 -20 = 60
ing_repo.set_stock(ing.id, 60)

passed = []
failed = []

def check(name, cond, extra=""):
    if cond:
        passed.append(name)
        print("PASS:", name)
    else:
        failed.append(name)
        print("FAIL:", name, extra)

import openpyxl
tmp = tempfile.mkdtemp()

# ===== Fix 1: Daily stock sheet =====
out_daily = os.path.join(tmp, "daily.xlsx")
ReportGenerator.export_daily_stock_sheet(out_daily, in_repo, out_repo, 2026, 6, 15)
wb = openpyxl.load_workbook(out_daily)
ws = wb["出入库表"]
check("daily D2 date", ws["D2"].value == datetime(2026, 6, 15), str(ws["D2"].value))
check("daily in row cat", ws.cell(row=5, column=1).value == "Veg", str(ws.cell(row=5, column=1).value))
check("daily in row name", ws.cell(row=5, column=2).value == "Cabbage", str(ws.cell(row=5, column=2).value))
check("daily in qty aggregated", ws.cell(row=5, column=3).value == 80, str(ws.cell(row=5, column=3).value))
check("daily in amount aggregated", ws.cell(row=5, column=6).value == 190, str(ws.cell(row=5, column=6).value))
check("daily out qty", ws.cell(row=5, column=12).value == 20, str(ws.cell(row=5, column=12).value))
check("daily out amount", ws.cell(row=5, column=15).value == 50, str(ws.cell(row=5, column=15).value))

# invalid day must raise
raised = False
try:
    ReportGenerator.export_daily_stock_sheet(os.path.join(tmp, "x.xlsx"),
                                             in_repo, out_repo, 2026, 6, 31)
except Exception as e:
    raised = "超出" in str(e)
check("daily invalid day raises", raised)

# ===== Fix 1: Monthly summary =====
out_month = os.path.join(tmp, "month.xlsx")
ReportGenerator.export_monthly_summary(out_month, in_repo, out_repo, 2026, 6)
wb = openpyxl.load_workbook(out_month)
ws_in = wb["2.每月采购食材汇总表"]
ws_out = wb["4.每月发出食材汇总表"]
check("monthly in name", ws_in.cell(row=4, column=2).value == "Cabbage", str(ws_in.cell(row=4, column=2).value))
check("monthly in qty", ws_in.cell(row=4, column=4).value == 80, str(ws_in.cell(row=4, column=4).value))
check("monthly in amount", ws_in.cell(row=4, column=6).value == 190, str(ws_in.cell(row=4, column=6).value))
check("monthly out qty", ws_out.cell(row=4, column=4).value == 20, str(ws_out.cell(row=4, column=4).value))
check("monthly out amount", ws_out.cell(row=4, column=6).value == 50, str(ws_out.cell(row=4, column=6).value))

# ===== Fix 1: Financial report =====
out_fin = os.path.join(tmp, "fin.xlsx")
ReportGenerator.export_financial_report(out_fin, in_repo, out_repo, 2026)
wb = openpyxl.load_workbook(out_fin)
ws_june = None
for name in wb.sheetnames:
    if "6月" in name.replace(" ", ""):
        ws_june = wb[name]
        break
check("financial june sheet exists", ws_june is not None)
if ws_june is not None:
    check("financial B5 filled", ws_june["B5"].value == 190, str(ws_june["B5"].value))
    check("financial E5 filled", ws_june["E5"].value == 190, str(ws_june["E5"].value))

# ===== Fix 1: Inspection report =====
out_insp = os.path.join(tmp, "insp.xlsx")
ReportGenerator.export_inspection_report(out_insp, insp_repo, 2026, 6)
wb = openpyxl.load_workbook(out_insp)
ws = wb["Sheet1"]
check("insp name", ws.cell(row=4, column=2).value == "Cabbage", str(ws.cell(row=4, column=2).value))
check("insp unit", ws.cell(row=4, column=3).value == "jin", str(ws.cell(row=4, column=3).value))
check("insp qty", ws.cell(row=4, column=4).value == 50, str(ws.cell(row=4, column=4).value))
check("insp result", ws.cell(row=4, column=14).value == "OK", str(ws.cell(row=4, column=14).value))
check("insp inspector", ws.cell(row=4, column=16).value == "Zhang", str(ws.cell(row=4, column=16).value))

# ===== Fix 2: Delete rollback =====
# current stock after seed: 0 +50 +30 -20 = 60
check("stock before del = 60", ing_repo.get_stock(ing.id) == 60, str(ing_repo.get_stock(ing.id)))
svc.delete_stock_in(si1.id)
check("stock after del in = 10", ing_repo.get_stock(ing.id) == 10, str(ing_repo.get_stock(ing.id)))
check("si1 gone", StockIn.get_or_none(StockIn.id == si1.id) is None)
so = StockOut.select().where(StockOut.ingredient == ing.id).first()
svc.delete_stock_out(so.id)
check("stock after del out = 30", ing_repo.get_stock(ing.id) == 30, str(ing_repo.get_stock(ing.id)))
check("so gone", StockOut.get_or_none(StockOut.id == so.id) is None)

# ===== Fix 3: Permission gating on created_at =====
class FakeUser:
    def __init__(self, role, uid=1):
        self.role = role
        self.id = uid

before = ing_repo.get_stock(ing.id)
Session.current_user = FakeUser("user")
rec = svc.stock_in(ing.id, 10, 1.0, created_at="2020-01-01 00:00:00")
check("normal user time ignored", rec.created_at.year == datetime.now().year,
      str(rec.created_at))
check("normal user stock +10", ing_repo.get_stock(ing.id) == before + 10, str(ing_repo.get_stock(ing.id)))
svc.delete_stock_in(rec.id)

Session.current_user = FakeUser("manager")
rec2 = svc.stock_in(ing.id, 5, 1.0, created_at="2020-01-01 00:00:00")
check("manager time honored", rec2.created_at == datetime(2020, 1, 1, 0, 0, 0),
      str(rec2.created_at))
svc.delete_stock_in(rec2.id)
Session.current_user = None

# ===== Fix 4: Data cleanup must not crash (sqlite_sequence / order) =====
repo = ReportRepository()
StockIn.create(ingredient=ing, quantity=5, unit_price=1, total_price=5)
crash = False
try:
    repo.clear_stock_records()
    repo.reset_autoincrement()
except Exception as e:
    crash = True
    print("  cleanup error:", repr(e))
check("cleanup no crash", not crash)
check("stock_in cleared", StockIn.select().count() == 0, str(StockIn.select().count()))
check("stock_out cleared", StockOut.select().count() == 0, str(StockOut.select().count()))
check("inspection cleared", InspectionRecord.select().count() == 0, str(InspectionRecord.select().count()))

# ===== Fix 6: FK on delete user (OperationLog SET NULL) =====
del_crash = False
try:
    User.delete().where(User.id == admin.id).execute()
except Exception as e:
    del_crash = True
    print("  user-delete error:", repr(e))
check("delete user no crash", not del_crash)
log_rec = OperationLog.select().where(OperationLog.action == "stock_in").first()
check("log survives user delete", log_rec is not None)
check("log user set null", log_rec is not None and log_rec.user is None)

print()
print("=== RESULT: %d passed, %d failed ===" % (len(passed), len(failed)))
if failed:
    print("FAILED:", failed)
    sys.exit(1)
print("ALL CHECKS PASSED")

"""生成食材导入模板"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MAC_BLUE = "0071e3"
MAC_GREEN = "34c759"
MAC_RED = "ff3b30"
MAC_ORANGE = "ff9500"
MAC_GRAY = "86868b"
MAC_LIGHT_GRAY = "f5f5f7"
MAC_BORDER = "d2d2d7"

def create_template():
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '食材导入模板.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "食材导入模板"

    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "学校食堂食材导入模板"
    title_cell.font = Font(size=18, bold=True, color=MAC_BLUE, name="Microsoft YaHei")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:F2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = "请按照下方格式填写食材信息，带 * 号为必填项"
    subtitle_cell.font = Font(size=11, color=MAC_GRAY, name="Microsoft YaHei")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    headers = [
        ("食材名称*", "食材的完整名称，如：大白菜、五花肉"),
        ("分类", "所属分类，如：蔬菜类、肉类、粮油类"),
        ("规格", "规格描述，如：新鲜、精品、散装"),
        ("单位*", "计量单位，如：斤、个、袋、桶"),
        ("安全库存", "库存预警阈值，低于此值会提醒补货"),
        ("供应商", "供应商名称，不存在会自动创建")
    ]

    header_fill = PatternFill(start_color=MAC_LIGHT_GRAY, end_color=MAC_LIGHT_GRAY, fill_type="solid")
    header_font = Font(color="1d1d1f", bold=True, size=12, name="Microsoft YaHei")
    header_border = Border(bottom=Side(style='medium', color=MAC_BORDER))

    for col, (header, _) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[4].height = 28

    examples = [
        ["大白菜", "蔬菜类", "新鲜", "斤", 50, "绿源蔬菜批发"],
        ["五花肉", "肉类", "精品", "斤", 30, "鸿运肉业"],
        ["鸡蛋", "蛋类", "散养土鸡蛋", "个", 200, "阳光养殖场"],
        ["东北大米", "粮油类", "五常稻花香", "袋", 20, "金穗粮油"],
        ["食用油", "粮油类", "非转基因压榨", "桶", 10, "金穗粮油"],
        ["西红柿", "蔬菜类", "有机", "斤", 40, "绿源蔬菜批发"],
        ["鸡胸肉", "肉类", "冷冻", "斤", 25, "鸿运肉业"],
        ["豆腐", "豆制品", "嫩豆腐", "块", 30, "豆香坊"],
    ]

    cell_font = Font(size=11, name="Microsoft YaHei")
    for row_idx, example in enumerate(examples, 5):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = Border(bottom=Side(style='thin', color='f0f0f0'))
        ws.row_dimensions[row_idx].height = 24

    column_widths = [18, 14, 18, 10, 14, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws2 = wb.create_sheet("填写说明")
    ws2.merge_cells('A1:C1')
    ws2['A1'] = "填写说明"
    ws2['A1'].font = Font(size=16, bold=True, color=MAC_BLUE, name="Microsoft YaHei")
    ws2['A1'].alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 32

    ws2['A3'] = "字段名"
    ws2['B3'] = "说明"
    ws2['C3'] = "是否必填"
    for col in range(1, 4):
        cell = ws2.cell(row=3, column=col)
        cell.font = Font(bold=True, size=12, name="Microsoft YaHei")
        cell.fill = PatternFill(start_color=MAC_LIGHT_GRAY, end_color=MAC_LIGHT_GRAY, fill_type="solid")
        cell.border = Border(bottom=Side(style='medium', color=MAC_BORDER))

    field_details = [
        ("食材名称", "食材的完整名称，建议简洁明了", "是"),
        ("分类", "食材所属分类。常见分类：蔬菜类、肉类、水产类、豆制品、粮油类、调味品、蛋类、水果类。不存在的分类会自动创建", "否"),
        ("规格", "食材的规格描述，如产地、等级、包装方式等", "否"),
        ("单位", "计量单位，如：斤、公斤、个、袋、桶、箱等", "是"),
        ("安全库存", "库存预警阈值。当库存低于此值时系统会提醒补货。不填默认为0", "否"),
        ("供应商", "供应商名称。不存在的供应商会自动创建", "否"),
    ]

    for row_idx, (field, desc, required) in enumerate(field_details, 4):
        ws2.cell(row=row_idx, column=1, value=field).font = Font(bold=True, size=11, name="Microsoft YaHei")
        ws2.cell(row=row_idx, column=2, value=desc).font = Font(size=11, name="Microsoft YaHei")
        req_cell = ws2.cell(row=row_idx, column=3, value=required)
        req_cell.font = Font(bold=True, size=11, color=MAC_RED if required == "是" else MAC_GREEN, name="Microsoft YaHei")
        req_cell.alignment = Alignment(horizontal="center")
        ws2.row_dimensions[row_idx].height = 28

    notice_row = len(field_details) + 6
    ws2.merge_cells(f'A{notice_row}:C{notice_row}')
    ws2.cell(row=notice_row, column=1, value="注意事项").font = Font(size=14, bold=True, color=MAC_ORANGE, name="Microsoft YaHei")

    notices = [
        "1. 第一行（第4行）为表头，请勿删除或修改",
        "2. 从第5行开始填写实际数据",
        "3. 分类和供应商如果不存在，导入时会自动创建",
        "4. 安全库存请填写数字，不填默认为0",
        "5. 建议先下载此模板，按格式填写后再导入",
        "6. 导入前请确保数据准确，避免重复导入",
    ]

    for i, notice in enumerate(notices, notice_row + 1):
        ws2.merge_cells(f'A{i}:C{i}')
        cell = ws2.cell(row=i, column=1, value=notice)
        cell.font = Font(size=11, name="Microsoft YaHei")
        ws2.row_dimensions[i].height = 24

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 55
    ws2.column_dimensions['C'].width = 12

    ws3 = wb.create_sheet("分类参考")
    ws3['A1'] = "系统默认分类"
    ws3['A1'].font = Font(size=14, bold=True, color=MAC_BLUE, name="Microsoft YaHei")

    default_categories = ["蔬菜类", "肉类", "水产类", "豆制品", "粮油类", "调味品", "蛋类", "水果类"]
    for i, cat in enumerate(default_categories, 3):
        ws3.cell(row=i, column=1, value=cat).font = Font(size=11, name="Microsoft YaHei")
    ws3.column_dimensions['A'].width = 16

    wb.save(file_path)
    print(f"模板已生成: {file_path}")

if __name__ == '__main__':
    create_template()
